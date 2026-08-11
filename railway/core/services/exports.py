from __future__ import annotations

from collections import OrderedDict, defaultdict
from decimal import Decimal
from io import BytesIO

from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import (
    AccountType, AuditLog, BankTransaction, CardTransaction, Charge, LegalNotice, Member, MessageRecipient,
    MonthlyJob, Payment, PaymentAllocationLine, Prepayment, Refund, UploadedFile, Vehicle,
)
from core.services.ledger import member_balance_snapshot

HEADER_FILL = PatternFill('solid', fgColor='D9EAF7')
SUBTOTAL_FILL = PatternFill('solid', fgColor='FFF2CC')


def _effective_job_q(prefix=''):
    field = f'{prefix}monthly_job'
    return Q(**{f'{field}__isnull': True}) | Q(**{f'{field}__is_current': True})


def _write_table(ws, headers, rows, freeze='A2'):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    widths = [len(str(h)) for h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            if cell.value is not None:
                widths[idx] = min(max(widths[idx], len(str(cell.value))), 50)
            if isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '#,##0'
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(10, width + 2)


def _member_rows():
    for member in Member.objects.filter(is_active_record=True).order_by('name', 'id'):
        vehicle = member.current_vehicle
        yield [
            member.id, member.name, member.birth6, vehicle.vehicle_no if vehicle else '',
            vehicle.purpose_char if vehicle else '', member.get_membership_status_display(),
            member.get_operational_status_display(), member.phone, member.address,
            member.official_address, member.region, member.memo,
            member.membership_started_on, member.certificate_issued_on,
            member.address_needs_check, member.phone_needs_check, member.sms_opt_out,
            member.get_collection_status_display(),
        ]


def _receivable_rows(job: MonthlyJob | None = None):
    charges = Charge.objects.filter(status=Charge.Status.POSTED)
    if job:
        charges = charges.filter(Q(monthly_job=job) | Q(monthly_job__isnull=True, charge_date__lte=job.period_end))
    else:
        charges = charges.filter(Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True))
    grouped = OrderedDict()
    for charge in charges.select_related('member').order_by('member__name', 'account_type', 'charge_date'):
        key = (charge.member_id, charge.account_type)
        data = grouped.setdefault(key, {
            'member': charge.member, 'account': charge.get_account_type_display(),
            'charged': Decimal('0'), 'settled': Decimal('0'), 'balance': Decimal('0'),
            'oldest': None,
        })
        data['charged'] += charge.amount
        data['settled'] += charge.settled_amount
        data['balance'] += charge.balance
        if charge.balance > 0 and data['oldest'] is None:
            data['oldest'] = charge.charge_date
    for data in grouped.values():
        if data['balance'] <= 0:
            continue
        member = data['member']
        vehicle = member.current_vehicle
        yield [
            member.id, member.name, vehicle.vehicle_no if vehicle else '', data['account'],
            data['charged'], data['settled'], data['balance'], data['oldest'],
            member.phone, member.get_operational_status_display(), member.region,
        ]


def _payment_rows(job: MonthlyJob | None = None):
    payments = Payment.objects.filter(is_effective=True).exclude(status=Payment.Status.CANCELLED)
    if job:
        payments = payments.filter(monthly_job=job)
    else:
        payments = payments.filter(Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True))
    for payment in payments.select_related('bank_transaction', 'card_transaction').order_by('payment_date', 'id'):
        lines = list(payment.allocation_lines.filter(status=PaymentAllocationLine.Status.ACTIVE).select_related('member'))
        if not lines:
            yield [payment.id, payment.payment_date, payment.get_source_type_display(), payment.amount, '', '', '', payment.unallocated_amount, payment.get_status_display()]
        for line in lines:
            yield [
                payment.id, payment.payment_date, payment.get_source_type_display(), payment.amount,
                line.member.name, line.get_account_type_display(), line.amount,
                payment.unallocated_amount, payment.get_status_display(),
            ]


def _unmatched_rows(job: MonthlyJob | None = None):
    txs = BankTransaction.objects.exclude(status__in=[BankTransaction.Status.AUTO_MATCHED, BankTransaction.Status.MANUAL_MATCHED, BankTransaction.Status.IGNORED])
    if job:
        txs = txs.filter(job=job)
    else:
        txs = txs.filter(job__is_current=True)
    for tx in txs.select_related('uploaded_file').order_by('transaction_at', 'id'):
        yield [
            tx.id, tx.job.year, tx.job.month, tx.uploaded_file.get_slot_type_display(),
            tx.transaction_at, tx.payer_text, tx.amount, tx.get_status_display(),
            tx.match_reason, tx.source_sheet, tx.source_row,
        ]
    files = UploadedFile.objects.filter(import_issues__status='open').distinct()
    if job:
        files = files.filter(job=job)
    else:
        files = files.filter(job__is_current=True)
    for file in files:
        for issue in file.import_issues.filter(status='open'):
            yield [
                f'ISSUE-{issue.id}', file.job.year, file.job.month, file.get_slot_type_display(),
                '', '', '', issue.issue_type, issue.message, issue.sheet_name, issue.source_row,
            ]


def _prepayment_rows():
    for prep in Prepayment.objects.filter(balance__gt=0).select_related('member').order_by('member__name'):
        vehicle = prep.member.current_vehicle
        yield [
            prep.member.id, prep.member.name, vehicle.vehicle_no if vehicle else '',
            prep.get_account_type_display(), prep.balance, prep.member.get_operational_status_display(),
        ]


def _closure_refund_rows():
    for member in Member.objects.filter(operational_status=Member.OperationalStatus.CLOSED).order_by('name'):
        refunds = list(member.refunds.order_by('-created_at'))
        if not refunds:
            yield [member.id, member.name, member.closed_on, '', '', '', '', '', '']
        for refund in refunds:
            yield [
                member.id, member.name, member.closed_on, refund.get_account_type_display(), refund.amount,
                refund.get_status_display(), refund.bank, refund.account_no, refund.holder,
                refund.refund_date, refund.method, refund.memo,
            ]


def _communication_rows():
    for recipient in MessageRecipient.objects.select_related('batch', 'member').order_by('-created_at'):
        yield [
            '문자', recipient.member.name, recipient.batch.get_message_type_display(),
            recipient.phone, recipient.amount_snapshot, recipient.body,
            recipient.get_status_display(), recipient.sent_at, recipient.failure_reason or recipient.exclusion_reason,
        ]
    for notice in LegalNotice.objects.select_related('member').order_by('-sent_date'):
        yield [
            '내용증명', notice.member.name, notice.get_address_type_display(),
            notice.registered_no, '', notice.address_snapshot,
            notice.get_delivery_status_display(), notice.sent_date, notice.memo,
        ]


def _card_rows(job: MonthlyJob | None = None):
    txs = CardTransaction.objects.filter(is_effective=True).select_related('payment')
    if job:
        txs = txs.filter(job=job)
    else:
        txs = txs.filter(job__is_current=True)
    for tx in txs.order_by('provider', 'transaction_at', 'id'):
        allocated = tx.payment.allocated_amount if hasattr(tx, 'payment') else Decimal('0')
        yield [
            tx.id, tx.get_provider_display(), tx.transaction_at, tx.member_name,
            tx.vehicle_no, tx.gross_amount, tx.fee_amount, tx.net_amount,
            tx.settlement_date, allocated, tx.get_status_display(), tx.txn_key,
        ]


def _bank_raw_rows(txs):
    txs = list(txs)
    raw_headers = []
    seen = set()
    for tx in txs:
        for key in tx.raw_data.keys():
            if key not in seen:
                raw_headers.append(key)
                seen.add(key)
    headers = raw_headers + ['시스템 거래ID', '처리상태', '매칭근거', '배정액', '미배정액', '카드순정산여부']
    rows = []
    for tx in txs:
        raw = [tx.raw_data.get(h, '') for h in raw_headers]
        rows.append(raw + [
            tx.id, tx.get_status_display(), tx.match_reason, tx.allocated_amount,
            tx.unallocated_amount, tx.is_card_settlement,
        ])
    return headers, rows


def build_workbook(*, job: MonthlyJob | None = None) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('미수금 현황')
    _write_table(ws, ['회원ID', '성명', '차량번호', '계정', '부과총액', '충당액', '미수잔액', '최초미수일', '휴대전화', '상태', '지역'], _receivable_rows(job))

    ws = wb.create_sheet('입금 처리내역')
    _write_table(ws, ['입금ID', '입금일시', '원천', '원입금액', '회원', '계정', '배정액', '미배정액', '상태'], _payment_rows(job))

    ws = wb.create_sheet('미배정·확인필요')
    _write_table(ws, ['거래/이슈ID', '연도', '월', '원본파일', '거래일시', '입금자', '금액', '상태/유형', '사유', '시트', '원본행'], _unmatched_rows(job))

    ws = wb.create_sheet('선납금 현황')
    _write_table(ws, ['회원ID', '성명', '차량번호', '계정', '선납잔액', '회원상태'], _prepayment_rows())

    ws = wb.create_sheet('폐업·환불대기')
    _write_table(ws, ['회원ID', '성명', '폐업일', '계정', '환불금액', '상태', '은행', '계좌번호', '예금주', '환불일', '방법', '메모'], _closure_refund_rows())

    ws = wb.create_sheet('회원 기본정보')
    _write_table(ws, ['회원ID', '성명', '주민번호앞6', '차량번호', '용도기호', '가입상태', '운영상태', '휴대전화', '주소', '공문주소', '지역', '비고', '가입일', '자격증명발급일', '주소확인필요', '연락처확인필요', '수신거부', '미수연락상태'], _member_rows())

    tx_base = BankTransaction.objects.all().select_related('payment')
    if job:
        tx_base = tx_base.filter(job=job)
        slots = [UploadedFile.SlotType.BANK_1, UploadedFile.SlotType.BANK_2, UploadedFile.SlotType.BANK_3]
        for idx, slot in enumerate(slots, 1):
            headers, rows = _bank_raw_rows(tx_base.filter(uploaded_file__slot_type=slot).order_by('source_row'))
            ws = wb.create_sheet(f'통장 원본 {idx}')
            _write_table(ws, headers, rows)
    else:
        headers, rows = _bank_raw_rows(tx_base.filter(job__is_current=True).order_by('job__year', 'job__month', 'transaction_at', 'id'))
        ws = wb.create_sheet('전체 통장 원본')
        _write_table(ws, headers, rows)

    ws = wb.create_sheet('카드 결제·정산')
    _write_table(ws, ['카드거래ID', '업체', '결제일시', '성명', '차량번호', '총결제액', '수수료', '순정산액', '정산일', '회원배정액', '상태', '중복키'], _card_rows(job))

    ws = wb.create_sheet('문자·내용증명 이력')
    _write_table(ws, ['구분', '성명', '유형', '전화/등기번호', '금액', '문구/주소', '결과', '발송일시', '사유/메모'], _communication_rows())

    ws = wb.create_sheet('요약')
    title = str(job) if job else '전체 누적 현황'
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A3'] = '생성 기준'
    ws['B3'] = '다운로드 시점 최신 데이터'
    ws['A4'] = '회원 수'
    ws['B4'] = Member.objects.filter(is_active_record=True).count()
    ws['A5'] = '총 미수금'
    ws['B5'] = sum((m.total_outstanding for m in Member.objects.filter(is_active_record=True)), Decimal('0'))
    ws['A6'] = '총 선납금'
    ws['B6'] = sum(Prepayment.objects.values_list('balance', flat=True), Decimal('0'))
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws['B5'].number_format = '#,##0'
    ws['B6'].number_format = '#,##0'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _autosize_sheet(ws, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), max_width)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color='374151')
        cell.fill = PatternFill('solid', fgColor='EEF5F8')
        cell.alignment = Alignment(horizontal='center', vertical='center')


def build_bank_ledger_workbook(job: MonthlyJob) -> BytesIO:
    """Create the daily bank ledger with K:O fixed as region/vehicle/name/account/amount."""
    wb = Workbook()
    ws = wb.active
    ws.title = '2026년 통장'
    headers = [
        '거래일시', '입금계좌', '입금자명', '원입금액', '거래상태',
        '매칭근거', '원본파일', '원본행', '시스템ID', '미배정액',
        '지역', '차량번호', '이름', '계정', '금액',
    ]
    ws.append(headers)
    txs = BankTransaction.objects.filter(job=job, is_effective=True).select_related('payment', 'uploaded_file').order_by('transaction_at', 'id')
    for tx in txs:
        lines = list(tx.payment.allocation_lines.filter(status=PaymentAllocationLine.Status.ACTIVE).select_related('member')) if hasattr(tx, 'payment') else []
        if not lines:
            ws.append([
                tx.transaction_at, tx.bank_account_label, tx.payer_text, tx.amount,
                tx.get_status_display(), tx.match_reason, tx.uploaded_file.original_name,
                tx.source_row, tx.id, tx.unallocated_amount,
                '', '', '', AccountType.SUSPENSE.label, tx.unallocated_amount,
            ])
            continue
        first = True
        for line in lines:
            vehicle = line.member.current_vehicle
            ws.append([
                tx.transaction_at if first else '', tx.bank_account_label if first else '',
                tx.payer_text if first else '', tx.amount if first else '',
                tx.get_status_display() if first else '', tx.match_reason if first else '',
                tx.uploaded_file.original_name if first else '', tx.source_row if first else '',
                tx.id, tx.unallocated_amount if first else '',
                line.member.region, vehicle.vehicle_no if vehicle else '', line.member.name,
                line.get_account_type_display(), line.amount,
            ])
            first = False

    manual_payments = Payment.objects.filter(
        monthly_job=job, source_type=Payment.SourceType.MANUAL,
        is_effective=True,
    ).exclude(status=Payment.Status.CANCELLED).order_by('payment_date', 'id')
    for payment in manual_payments:
        lines = list(payment.allocation_lines.filter(
            status=PaymentAllocationLine.Status.ACTIVE
        ).select_related('member'))
        if not lines:
            continue
        first = True
        for line in lines:
            vehicle = line.member.current_vehicle
            ws.append([
                payment.payment_date if first else '', '수기입력' if first else '',
                payment.memo if first else '', payment.amount if first else '',
                payment.get_status_display() if first else '', '회원명단 수기입금' if first else '',
                '수기입력' if first else '', '', f'M-{payment.id}',
                payment.unallocated_amount if first else '',
                line.member.region, vehicle.vehicle_no if vehicle else '', line.member.name,
                line.get_account_type_display(), line.amount,
            ])
            first = False
    _style_header(ws)
    for col in ('D', 'J', 'O'):
        for cell in ws[col][1:]:
            cell.number_format = '#,##0'
    _autosize_sheet(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _monthly_payment_summary(member: Member, job: MonthlyJob):
    lines = PaymentAllocationLine.objects.filter(
        member=member,
        status=PaymentAllocationLine.Status.ACTIVE,
        payment__is_effective=True,
        payment__payment_date__date__gte=job.period_start,
        payment__payment_date__date__lte=job.period_end,
    ).select_related('payment').order_by('payment__payment_date', 'id')
    amount = sum((line.amount for line in lines), Decimal('0'))
    dates = [line.payment.payment_date.strftime('%m.%d') for line in lines]
    return amount, ', '.join(dates)


def build_receivables_workbook(job: MonthlyJob) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = '현재명단'
    headers = [
        '지역', '계정', '비고', '이름', '차량번호', '상태', '최초부과기준',
        f'{job.month}월 입금액', f'{job.month}월 입금날짜', '현재 잔액', '전화번호', '주소',
    ]
    ws.append(headers)
    members = list(Member.objects.filter(
        is_active_record=True,
        operational_status=Member.OperationalStatus.ACTIVE,
    ).order_by('region', 'name'))
    member_ids = [member.id for member in members]
    balances = member_balance_snapshot(member_ids)
    vehicle_map = {
        row['member_id']: row['vehicle_no']
        for row in Vehicle.objects.filter(member_id__in=member_ids, is_current=True).values('member_id', 'vehicle_no')
    }
    monthly_map = defaultdict(lambda: {'amount': Decimal('0'), 'dates': []})
    monthly_lines = PaymentAllocationLine.objects.filter(
        member_id__in=member_ids, status=PaymentAllocationLine.Status.ACTIVE,
        payment__is_effective=True,
        payment__payment_date__date__gte=job.period_start,
        payment__payment_date__date__lte=job.period_end,
    ).select_related('payment').order_by('payment__payment_date', 'id')
    for line in monthly_lines:
        monthly_map[line.member_id]['amount'] += line.amount
        monthly_map[line.member_id]['dates'].append(line.payment.payment_date.strftime('%m.%d'))
    for member in members:
        summary = monthly_map[member.id]
        account = member.get_receivable_account_type_display() if member.receivable_account_type else (
            '협회비' if member.membership_status == Member.MembershipStatus.ACTIVE else '관리비'
        )
        anchor = member.membership_started_on if account == '협회비' else member.certificate_issued_on
        ws.append([
            member.region, account, member.memo, member.name, vehicle_map.get(member.id, ''),
            member.get_operational_status_display(), anchor, summary['amount'], ', '.join(summary['dates']),
            balances.get(member.id, {}).get('net_balance', Decimal('0')),
            member.phone, member.address,
        ])
    _style_header(ws)
    for col in ('H', 'J'):
        for cell in ws[col][1:]:
            cell.number_format = '#,##0'
    _autosize_sheet(ws)

    closed = wb.create_sheet('폐업명단')
    closed.append(['지역', '계정', '비고', '이름', '차량번호', '폐업일', '현재 잔액', '환불대기', '전화번호', '주소'])
    closed_members = list(Member.objects.filter(
        is_active_record=True,
        operational_status=Member.OperationalStatus.CLOSED,
    ).order_by('-closed_on', 'name'))
    closed_ids = [member.id for member in closed_members]
    closed_balances = member_balance_snapshot(closed_ids)
    closed_vehicle_map = {
        row['member_id']: row['vehicle_no']
        for row in Vehicle.objects.filter(member_id__in=closed_ids, is_current=True).values('member_id', 'vehicle_no')
    }
    refund_map = {
        row['member_id']: row['total'] or Decimal('0')
        for row in Refund.objects.filter(member_id__in=closed_ids, status=Refund.Status.PENDING).values('member_id').annotate(total=Sum('amount'))
    }
    for member in closed_members:
        account = member.get_receivable_account_type_display() if member.receivable_account_type else (
            '협회비' if member.membership_status == Member.MembershipStatus.ACTIVE else '관리비'
        )
        closed.append([
            member.region, account, member.memo, member.name, closed_vehicle_map.get(member.id, ''),
            member.closed_on, closed_balances.get(member.id, {}).get('net_balance', Decimal('0')),
            refund_map.get(member.id, Decimal('0')), member.phone, member.address,
        ])
    _style_header(closed)
    for col in ('G', 'H'):
        for cell in closed[col][1:]:
            cell.number_format = '#,##0'
    _autosize_sheet(closed)

    detail = wb.create_sheet('입금일자 상세')
    detail.append(['입금일시', '원천', '입금자명', '회원', '차량번호', '계정', '반영금액', '상태'])
    lines = PaymentAllocationLine.objects.filter(
        status=PaymentAllocationLine.Status.ACTIVE,
        payment__is_effective=True,
        payment__payment_date__date__gte=job.period_start,
        payment__payment_date__date__lte=job.period_end,
    ).select_related('payment', 'member', 'payment__bank_transaction', 'payment__card_transaction').order_by('payment__payment_date', 'id')
    for line in lines:
        payment = line.payment
        if payment.bank_transaction_id:
            payer = payment.bank_transaction.payer_text
        elif payment.card_transaction_id:
            payer = payment.card_transaction.member_name
        else:
            payer = payment.memo
        vehicle = line.member.current_vehicle
        detail.append([
            payment.payment_date, payment.get_source_type_display(), payer, line.member.name,
            vehicle.vehicle_no if vehicle else '', line.get_account_type_display(), line.amount,
            payment.get_status_display(),
        ])
    _style_header(detail)
    for cell in detail['G'][1:]:
        cell.number_format = '#,##0'
    _autosize_sheet(detail)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_voucher_workbook(job: MonthlyJob) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = '입금전표'
    headers = [
        '전표번호', '입금일자', '입금계좌', '입금자명', '지역', '차량번호', '회원명',
        '차변계정', '대변계정', '금액', '적요', '처리상태', '원본거래ID',
    ]
    ws.append(headers)
    txs = BankTransaction.objects.filter(job=job, is_effective=True).select_related('payment').order_by('transaction_at', 'id')
    daily_seq = defaultdict(int)
    for tx in txs:
        payment_date = (tx.transaction_at or timezone.now()).date()
        daily_seq[payment_date] += 1
        voucher_no = f'{payment_date:%Y%m%d}-{daily_seq[payment_date]:03d}'
        lines = list(tx.payment.allocation_lines.filter(status=PaymentAllocationLine.Status.ACTIVE).select_related('member')) if hasattr(tx, 'payment') else []
        if not lines:
            ws.append([
                voucher_no, payment_date, tx.bank_account_label, tx.payer_text, '', '', '',
                '보통예금', '가수금', tx.amount, tx.payer_text, tx.get_status_display(), tx.id,
            ])
            continue
        allocated_total = Decimal('0')
        for line in lines:
            allocated_total += line.amount
            vehicle = line.member.current_vehicle
            ws.append([
                voucher_no, payment_date, tx.bank_account_label, tx.payer_text, line.member.region,
                vehicle.vehicle_no if vehicle else '', line.member.name, '보통예금',
                line.get_account_type_display(), line.amount,
                f'{vehicle.vehicle_no if vehicle else ""} {line.member.name} {line.get_account_type_display()}'.strip(),
                tx.get_status_display(), tx.id,
            ])
        if tx.amount > allocated_total:
            ws.append([
                voucher_no, payment_date, tx.bank_account_label, tx.payer_text, '', '', '',
                '보통예금', '가수금', tx.amount - allocated_total, '미배정 잔액', '확인 필요', tx.id,
            ])

    manual_payments = Payment.objects.filter(
        monthly_job=job, source_type=Payment.SourceType.MANUAL,
        is_effective=True,
    ).exclude(status=Payment.Status.CANCELLED).order_by('payment_date', 'id')
    for payment in manual_payments:
        payment_date = payment.payment_date.date()
        daily_seq[payment_date] += 1
        voucher_no = f'{payment_date:%Y%m%d}-{daily_seq[payment_date]:03d}'
        for line in payment.allocation_lines.filter(
            status=PaymentAllocationLine.Status.ACTIVE
        ).select_related('member'):
            vehicle = line.member.current_vehicle
            ws.append([
                voucher_no, payment_date, '수기입력', payment.memo, line.member.region,
                vehicle.vehicle_no if vehicle else '', line.member.name, '보통예금',
                line.get_account_type_display(), line.amount,
                f'{vehicle.vehicle_no if vehicle else ""} {line.member.name} {line.get_account_type_display()}'.strip(),
                payment.get_status_display(), f'M-{payment.id}',
            ])
    _style_header(ws)
    for cell in ws['J'][1:]:
        cell.number_format = '#,##0'
    _autosize_sheet(ws)

    summary = wb.create_sheet('계정별 합계')
    summary.append(['계정', '금액'])
    account_totals = defaultdict(Decimal)
    for row in ws.iter_rows(min_row=2, values_only=True):
        account_totals[row[8]] += Decimal(str(row[9] or 0))
    for account, amount in sorted(account_totals.items()):
        summary.append([account, amount])
    _style_header(summary)
    for cell in summary['B'][1:]:
        cell.number_format = '#,##0'
    _autosize_sheet(summary)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

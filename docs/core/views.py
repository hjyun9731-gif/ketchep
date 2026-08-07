from __future__ import annotations

from datetime import datetime, time
from decimal import Decimal
from io import BytesIO
import zipfile

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.core.paginator import Paginator
from django.core.cache import cache
from django.db.models import Count, DecimalField, ExpressionWrapper, F, OuterRef, Q, Subquery, Sum, Value
from django.db.models.functions import Coalesce
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from core.forms import (
    AllocationFormSet, ArrearsComposeForm, BankPasteForm, CloseMemberForm, ColumnMappingForm,
    InitialDataImportForm,
    JoinAssociationForm, LeaveAssociationForm, LegalNoticeForm, ManualPaymentForm, MemberForm, QuickMemberForm,
    MessageScheduleEditForm, MonthlyJobForm, RefundForm, RefundMessageForm, RefundPendingForm,
    ReopenMemberForm, SimpleExcelUploadForm, TransferMemberForm, UploadForm,
)
from core.models import (
    AccountType, AuditLog, BankTransaction, CardTransaction, Charge, ChargeSettlement, HistoricalPaymentRecord, ImportIssue, LegalNotice, Member, PayerAlias,
    MessageBatch, MessageRecipient, MessageTemplate, MonthlyJob, Payment, MemberLink,
    PaymentAllocationLine, Prepayment, Refund, UploadedFile, Vehicle,
)
from core.services.audit import log_action
from core.services.billing import (
    close_member, generate_charges_for_job, generate_due_charges_through_today,
    join_association, leave_association, reopen_member,
)
from core.services.exports import (
    build_bank_ledger_workbook, build_receivables_workbook, build_voucher_workbook, build_workbook,
)
from core.services.imports import parse_uploaded_file, process_uploaded_file
from core.services.historical_payments import backfill_receivable_payment_history
from core.services.paste_import import get_or_create_current_job, process_pasted_bank_text
from core.services.jobs import create_job_version, clone_latest_uploaded_files, set_current_job
from core.services.ledger import complete_refund, replace_payment_allocations
from core.services.matching import (
    auto_match_bank_job, auto_match_card_job, copy_allocations_from_previous_version,
)
from core.services.transfers import transfer_member
from core.services.messaging import (
    BalsongClient,
    cancel_batch,
    create_arrears_batch,
    create_refund_batch,
    retry_failed_batch,
    send_batch,
    sync_batch_results,
    update_batch_schedule,
)
from core.utils import json_safe_model, normalize_text, normalize_vehicle_no, sha256_file


MONEY_FIELD = DecimalField(max_digits=14, decimal_places=2)
MONEY_ZERO_VALUE = Value(Decimal('0.00'), output_field=MONEY_FIELD)


def _member_queryset_with_financials(queryset):
    """Attach current vehicle and outstanding balance without per-member queries."""
    current_vehicle = (
        Vehicle.objects.filter(member_id=OuterRef('pk'), is_current=True)
        .order_by('-start_date', '-id')
        .values('vehicle_no')[:1]
    )
    charge_total = (
        Charge.objects.filter(
            member_id=OuterRef('pk'),
            status=Charge.Status.POSTED,
        )
        .filter(Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True))
        .values('member_id')
        .annotate(total=Sum('amount'))
        .values('total')[:1]
    )
    settlement_total = (
        ChargeSettlement.objects.filter(
            charge__member_id=OuterRef('pk'),
            charge__status=Charge.Status.POSTED,
            is_active=True,
        )
        .filter(Q(charge__monthly_job__isnull=True) | Q(charge__monthly_job__is_current=True))
        .values('charge__member_id')
        .annotate(total=Sum('amount'))
        .values('total')[:1]
    )
    prepayment_total = (
        Prepayment.objects.filter(member_id=OuterRef('pk'), balance__gt=0)
        .values('member_id')
        .annotate(total=Sum('balance'))
        .values('total')[:1]
    )
    return (
        queryset.annotate(
            current_vehicle_no=Subquery(current_vehicle),
            charge_total_fast=Coalesce(
                Subquery(charge_total, output_field=MONEY_FIELD),
                MONEY_ZERO_VALUE,
            ),
            settlement_total_fast=Coalesce(
                Subquery(settlement_total, output_field=MONEY_FIELD),
                MONEY_ZERO_VALUE,
            ),
            prepayment_amount=Coalesce(
                Subquery(prepayment_total, output_field=MONEY_FIELD),
                MONEY_ZERO_VALUE,
            ),
        )
        .annotate(
            outstanding_amount=ExpressionWrapper(
                F('charge_total_fast') - F('settlement_total_fast'),
                output_field=MONEY_FIELD,
            )
        )
    )


def _attach_member_page_data(page_obj):
    """Load vehicle and balance data only for the current page.

    The old list annotated and sorted all 3,000+ members with correlated
    subqueries before pagination.  That made even page 1 wait for the whole
    ledger.  This keeps pagination on indexed Member columns, then performs
    three compact aggregate queries for the 50 visible rows.
    """
    members = list(page_obj.object_list)
    member_ids = [member.id for member in members]
    if not member_ids:
        page_obj.object_list = members
        return page_obj

    vehicle_map = {
        row['member_id']: row['vehicle_no']
        for row in Vehicle.objects.filter(member_id__in=member_ids, is_current=True)
        .values('member_id', 'vehicle_no')
    }
    charge_map = {
        row['member_id']: row['total'] or Decimal('0')
        for row in Charge.objects.filter(
            member_id__in=member_ids, status=Charge.Status.POSTED,
        ).filter(
            Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True)
        ).values('member_id').annotate(total=Sum('amount'))
    }
    settlement_map = {
        row['charge__member_id']: row['total'] or Decimal('0')
        for row in ChargeSettlement.objects.filter(
            charge__member_id__in=member_ids, charge__status=Charge.Status.POSTED, is_active=True,
        ).filter(
            Q(charge__monthly_job__isnull=True) | Q(charge__monthly_job__is_current=True)
        ).values('charge__member_id').annotate(total=Sum('amount'))
    }
    prepayment_map = {
        row['member_id']: row['total'] or Decimal('0')
        for row in Prepayment.objects.filter(member_id__in=member_ids, balance__gt=0)
        .values('member_id').annotate(total=Sum('balance'))
    }
    for member in members:
        member.current_vehicle_no = vehicle_map.get(member.id, '')
        member.outstanding_amount = max(
            Decimal('0'),
            charge_map.get(member.id, Decimal('0')) - settlement_map.get(member.id, Decimal('0')),
        )
        member.prepayment_amount = max(Decimal('0'), prepayment_map.get(member.id, Decimal('0')))
        member.net_balance = member.outstanding_amount - member.prepayment_amount
    page_obj.object_list = members
    return page_obj


def _member_status_counts():
    cache_key = 'member-status-counts-v4'
    cached = cache.get(cache_key)
    if cached is not None:
        return cached
    value = Member.objects.filter(is_active_record=True).aggregate(
        active=Count('id', filter=Q(operational_status=Member.OperationalStatus.ACTIVE)),
        closed=Count('id', filter=Q(operational_status=Member.OperationalStatus.CLOSED)),
    )
    cache.set(cache_key, value, 300)
    return value


def _member_regions():
    cache_key = 'member-regions-v4'
    cached = cache.get(cache_key)
    if cached is not None:
        return cached
    value = list(
        Member.objects.filter(is_active_record=True)
        .exclude(region='')
        .values_list('region', flat=True)
        .distinct()
        .order_by('region')
    )
    cache.set(cache_key, value, 600)
    return value


def _invalidate_member_list_cache():
    cache.delete_many(['member-status-counts-v4', 'member-regions-v4'])


def _wants_modal(request):
    return request.GET.get('modal') == '1' or request.POST.get('modal') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest'


def _modal_form_response(request, template_name, context, *, status=200):
    return render(request, template_name, context, status=status)


def _actor(request):
    return request.user.username if request.user.is_authenticated else 'admin'


def _save_and_process_excel(*, job, slot_type, file_obj):
    digest = sha256_file(file_obj)
    existing = UploadedFile.objects.filter(
        job=job, slot_type=slot_type, sha256=digest,
        parse_status=UploadedFile.ParseStatus.PROCESSED,
    ).first()
    if existing:
        # A parser update must be able to repair data from the same original
        # workbook. Rebuild ParsedRow canonical values and apply them again.
        parse_uploaded_file(existing)
        if existing.parse_status == UploadedFile.ParseStatus.NEEDS_MAPPING:
            raise ValueError(existing.parse_error or '기존 파일의 열을 다시 찾지 못했습니다.')
        result = process_uploaded_file(existing)
        result = {**result, 'refreshed_existing_file': 1}
        return existing, result
    uploaded = UploadedFile.objects.create(
        job=job,
        slot_type=slot_type,
        file=file_obj,
        original_name=file_obj.name,
        sha256=digest,
        size=file_obj.size,
    )
    parse_uploaded_file(uploaded)
    if uploaded.parse_status == UploadedFile.ParseStatus.NEEDS_MAPPING:
        raise ValueError(
            uploaded.parse_error
            or f'{uploaded.get_slot_type_display()} 파일에서 필요한 열을 찾지 못했습니다.'
        )
    return uploaded, process_uploaded_file(uploaded)


@login_required
def dashboard(request):
    today = timezone.localdate()
    job = get_or_create_current_job(today)
    # Post only newly due individual charges once per day. The old full-member
    # recalculation on every dashboard request caused thousands of queries.
    generate_due_charges_through_today(job, actor=_actor(request))
    active_members = Member.objects.filter(
        is_active_record=True,
        operational_status=Member.OperationalStatus.ACTIVE,
    )
    closed_members = Member.objects.filter(
        is_active_record=True,
        operational_status=Member.OperationalStatus.CLOSED,
    )
    review_statuses = [
        BankTransaction.Status.UNMATCHED,
        BankTransaction.Status.REVIEW,
        BankTransaction.Status.DUPLICATE,
    ]
    current_transactions = BankTransaction.objects.filter(job=job, is_effective=True)
    today_transactions = current_transactions.filter(transaction_at__date=today)
    charge_total = Charge.objects.filter(
        member__in=active_members,
        status=Charge.Status.POSTED,
    ).filter(
        Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True)
    ).aggregate(v=Sum('amount'))['v'] or Decimal('0')
    settlement_total = ChargeSettlement.objects.filter(
        charge__member__in=active_members,
        charge__status=Charge.Status.POSTED,
        is_active=True,
    ).filter(
        Q(charge__monthly_job__isnull=True) | Q(charge__monthly_job__is_current=True)
    ).aggregate(v=Sum('amount'))['v'] or Decimal('0')
    summary = {
        'members': active_members.count(),
        'closed_members': closed_members.count(),
        'outstanding': charge_total - settlement_total,
        'today_amount': today_transactions.aggregate(v=Sum('amount'))['v'] or Decimal('0'),
        'today_count': today_transactions.count(),
        'review_transactions': current_transactions.filter(status__in=review_statuses).count(),
        'auto_matched': current_transactions.filter(status=BankTransaction.Status.AUTO_MATCHED).count(),
        'prepayment': Prepayment.objects.aggregate(v=Sum('balance'))['v'] or Decimal('0'),
    }
    recent = current_transactions.select_related('payment').order_by('-transaction_at', '-id')[:12]
    review_rows = current_transactions.filter(status__in=review_statuses).select_related('payment').order_by('-transaction_at', '-id')[:12]
    paste_forms = {
        UploadedFile.SlotType.BANK_1: BankPasteForm(initial={'slot_type': UploadedFile.SlotType.BANK_1}),
        UploadedFile.SlotType.BANK_2: BankPasteForm(initial={'slot_type': UploadedFile.SlotType.BANK_2}),
        UploadedFile.SlotType.BANK_3: BankPasteForm(initial={'slot_type': UploadedFile.SlotType.BANK_3}),
    }
    return render(request, 'core/dashboard.html', {
        'job': job,
        'summary': summary,
        'recent_transactions': recent,
        'review_rows': review_rows,
        'paste_forms': paste_forms,
        'today': today,
    })


@login_required
@transaction.atomic
def initial_data_import(request):
    form = InitialDataImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        job = get_or_create_current_job(timezone.localdate())
        try:
            license_upload, license_result = _save_and_process_excel(
                job=job,
                slot_type=UploadedFile.SlotType.LICENSE,
                file_obj=form.cleaned_data['license_file'],
            )
            receivable_upload, receivable_result = _save_and_process_excel(
                job=job,
                slot_type=UploadedFile.SlotType.RECEIVABLES,
                file_obj=form.cleaned_data['receivables_file'],
            )
            history_result = backfill_receivable_payment_history(receivable_upload)
            issue_count = ImportIssue.objects.filter(
                uploaded_file__in=[license_upload, receivable_upload],
                status=ImportIssue.Status.OPEN,
            ).count()
            messages.success(
                request,
                '기존 자료를 불러왔습니다. '
                f"회원 신규 {license_result.get('created_members', 0)}명 · "
                f"회원 갱신 {license_result.get('updated_members', 0)}명 · "
                f"기초 미수금 {receivable_result.get('opening_charges', 0)}건 · "
                f"선납금 {receivable_result.get('opening_prepayments', 0)}건 · "
                f"1~7월 입금이력 {history_result.get('created', 0) + history_result.get('updated', 0)}건 · "
                f'확인 필요 {issue_count}건',
            )
            return redirect('core:member_list')
        except Exception as exc:
            messages.error(request, f'기존 자료 불러오기 실패: {exc}')
    return render(request, 'core/initial_import.html', {'form': form})


@login_required
@transaction.atomic
def bank_paste(request):
    if request.method != 'POST':
        raise Http404
    form = BankPasteForm(request.POST)
    if not form.is_valid():
        messages.error(request, '붙여넣기 내용을 확인하세요.')
        return redirect('core:dashboard')
    try:
        result = process_pasted_bank_text(
            slot_type=form.cleaned_data['slot_type'],
            pasted_text=form.cleaned_data['pasted_text'],
            actor=_actor(request),
        )
        messages.success(
            request,
            '입금내역 처리 완료: '
            f"신규 {result['created_transactions']}건 · "
            f"자동매칭 {result['auto_matched']}건 · "
            f"확인필요 {result['review']}건 · "
            f"중복제외 {result['skipped_duplicates']}건",
        )
    except Exception as exc:
        messages.error(request, f'입금내역 처리 실패: {exc}')
    return redirect('core:dashboard')


@login_required
def job_create(request):
    if request.method == 'POST':
        form = MonthlyJobForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data['year']
            month = form.cleaned_data['month']
            based_on = MonthlyJob.objects.filter(year=year, month=month).order_by('-version').first()
            job = create_job_version(
                year=year, month=month,
                version_name=form.cleaned_data['version_name'],
                based_on=based_on,
                actor=_actor(request),
            )
            job.memo = form.cleaned_data.get('memo', '')
            job.save(update_fields=['memo', 'updated_at'])
            messages.success(request, '새 작업 버전을 만들었습니다.')
            return redirect('core:job_detail', pk=job.pk)
    else:
        now = timezone.localdate()
        form = MonthlyJobForm(initial={'year': now.year, 'month': now.month, 'version_name': '1차 작업'})
    return render(request, 'core/form.html', {'form': form, 'title': '월 작업 만들기'})


@login_required
def job_detail(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    latest_files = {}
    for file in job.uploaded_files.order_by('slot_type', '-created_at'):
        latest_files.setdefault(file.slot_type, file)
    bank_counts = {
        'total': job.bank_transactions.count(),
        'matched': job.bank_transactions.filter(status__in=[BankTransaction.Status.AUTO_MATCHED, BankTransaction.Status.MANUAL_MATCHED]).count(),
        'review': job.bank_transactions.filter(status__in=[BankTransaction.Status.REVIEW, BankTransaction.Status.DUPLICATE, BankTransaction.Status.UNMATCHED]).count(),
    }
    charge_counts = {
        'posted': job.charges.filter(status=Charge.Status.POSTED).count(),
        'cancelled': job.charges.filter(status=Charge.Status.CANCELLED).count(),
    }
    import_issues = ImportIssue.objects.filter(
        uploaded_file__job=job, status=ImportIssue.Status.OPEN,
    ).select_related('uploaded_file').order_by('uploaded_file__slot_type', 'source_row')[:100]
    return render(request, 'core/job_detail.html', {
        'job': job,
        'latest_files': latest_files,
        'slot_choices': UploadedFile.SlotType.choices,
        'bank_counts': bank_counts,
        'charge_counts': charge_counts,
        'import_issues': import_issues,
    })


@login_required
@transaction.atomic
def job_upload(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_obj = form.cleaned_data['file']
            slot_type = form.cleaned_data['slot_type']
            target_job = job
            if job.uploaded_files.filter(slot_type=slot_type).exists():
                next_version = (MonthlyJob.objects.filter(year=job.year, month=job.month).order_by('-version').first().version + 1)
                target_job = create_job_version(
                    year=job.year,
                    month=job.month,
                    version_name=f'{next_version}차 작업',
                    based_on=job,
                    actor=_actor(request),
                )
                clone_latest_uploaded_files(job, target_job, exclude_slot=slot_type)
                messages.info(request, f'같은 유형 파일이 있어 {target_job.version_name} 새 버전을 만들었습니다.')
            digest = sha256_file(file_obj)
            uploaded = UploadedFile.objects.create(
                job=target_job,
                slot_type=slot_type,
                file=file_obj,
                original_name=file_obj.name,
                sha256=digest,
                size=file_obj.size,
            )
            try:
                parse_uploaded_file(uploaded)
                if uploaded.parse_status == UploadedFile.ParseStatus.NEEDS_MAPPING:
                    messages.warning(request, '파일은 읽었지만 열 매핑이 필요합니다.')
                    return redirect('core:file_mapping', pk=uploaded.pk)
                messages.success(request, '파일을 업로드하고 파싱했습니다.')
            except Exception as exc:
                messages.error(request, f'파일 파싱 실패: {exc}')
            return redirect('core:job_detail', pk=target_job.pk)
    else:
        initial_slot = request.GET.get('slot')
        form = UploadForm(initial={'slot_type': initial_slot} if initial_slot else None)
    return render(request, 'core/form.html', {'form': form, 'title': f'{job} 파일 업로드'})


@login_required
def file_mapping(request, pk):
    uploaded = get_object_or_404(UploadedFile, pk=pk)
    if request.method == 'POST':
        form = ColumnMappingForm(request.POST, uploaded=uploaded)
        if form.is_valid():
            uploaded.column_mapping = {k: v for k, v in form.cleaned_data.items() if v}
            uploaded.parse_status = UploadedFile.ParseStatus.PARSED
            uploaded.parse_error = ''
            uploaded.save(update_fields=['column_mapping', 'parse_status', 'parse_error', 'updated_at'])
            parse_uploaded_file(uploaded)
            if uploaded.parse_status == UploadedFile.ParseStatus.PARSED:
                messages.success(request, '열 매핑을 저장했습니다.')
                return redirect('core:job_detail', pk=uploaded.job_id)
    else:
        form = ColumnMappingForm(uploaded=uploaded)
    sample_rows = uploaded.parsed_rows.all()[:5]
    return render(request, 'core/file_mapping.html', {'uploaded': uploaded, 'form': form, 'sample_rows': sample_rows})


@login_required
def file_process(request, pk):
    uploaded = get_object_or_404(UploadedFile, pk=pk)
    if request.method != 'POST':
        raise Http404
    try:
        result = process_uploaded_file(uploaded)
        messages.success(request, '파일을 업무원장에 반영했습니다: ' + ', '.join(f'{k}={v}' for k, v in result.items() if isinstance(v, (int, str))))
    except Exception as exc:
        messages.error(request, f'반영 실패: {exc}')
    return redirect('core:job_detail', pk=uploaded.job_id)


@login_required
def job_analyze(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    if request.method != 'POST':
        raise Http404
    errors = []
    priority = {
        UploadedFile.SlotType.LICENSE: 0,
        UploadedFile.SlotType.RECEIVABLES: 1,
        UploadedFile.SlotType.BANK_1: 2,
        UploadedFile.SlotType.BANK_2: 3,
        UploadedFile.SlotType.BANK_3: 4,
        UploadedFile.SlotType.ALTOLAN: 5,
        UploadedFile.SlotType.CIDER: 6,
    }
    parsed_files = list(job.uploaded_files.filter(parse_status=UploadedFile.ParseStatus.PARSED))
    parsed_files.sort(key=lambda item: priority.get(item.slot_type, 99))
    for uploaded in parsed_files:
        try:
            process_uploaded_file(uploaded)
        except Exception as exc:
            errors.append(f'{uploaded.original_name}: {exc}')
    try:
        charge_result = generate_charges_for_job(job, actor=_actor(request))
        copy_result = copy_allocations_from_previous_version(job)
        bank_result = auto_match_bank_job(job)
        card_result = auto_match_card_job(job)
        messages.success(request, f'분석 완료. 부과 {charge_result}, 이전매칭 {copy_result}, 통장 {bank_result}, 카드 {card_result}')
    except Exception as exc:
        errors.append(str(exc))
    for error in errors:
        messages.error(request, error)
    return redirect('core:job_detail', pk=job.pk)


@login_required
def job_generate_charges(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    if request.method == 'POST':
        try:
            result = generate_charges_for_job(job, actor=_actor(request))
            messages.success(request, f'부과 재계산 완료: {result}')
        except Exception as exc:
            messages.error(request, str(exc))
    return redirect('core:job_detail', pk=job.pk)


@login_required
def job_make_current(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    if request.method == 'POST':
        set_current_job(job, actor=_actor(request))
        messages.success(request, '현재 기준 버전으로 지정했습니다.')
    return redirect('core:job_detail', pk=pk)


@login_required
def job_finalize(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    if request.method == 'POST':
        unallocated = sum((tx.unallocated_amount for tx in job.bank_transactions.filter(is_effective=True)), Decimal('0'))
        review = job.bank_transactions.filter(status__in=[BankTransaction.Status.REVIEW, BankTransaction.Status.DUPLICATE]).count()
        if unallocated > 0 or review:
            messages.warning(request, f'미배정 {unallocated:,.0f}원, 확인필요 {review}건이 남아 있습니다. 확정은 가능하지만 다시 확인하세요.')
        job.status = MonthlyJob.Status.FINAL
        job.finalized_at = timezone.now()
        job.save(update_fields=['status', 'finalized_at', 'updated_at'])
        log_action(action='monthly_job_finalized', instance=job, actor=_actor(request))
        messages.success(request, '최종확정했습니다. 이후에도 수정할 수 있습니다.')
    return redirect('core:job_detail', pk=pk)


@login_required
def job_export(request, pk):
    job = get_object_or_404(MonthlyJob, pk=pk)
    bundle = BytesIO()
    with zipfile.ZipFile(bundle, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(f'{job.year}년 통장_완성본.xlsx', build_bank_ledger_workbook(job).getvalue())
        archive.writestr(f'{job.year}년 {job.month}월 미수금_완성본.xlsx', build_receivables_workbook(job).getvalue())
        archive.writestr(f'{job.year}년 {job.month}월 입금전표_완성본.xlsx', build_voucher_workbook(job).getvalue())
    bundle.seek(0)
    filename = f'{job.year}년_{job.month:02d}월_통장_미수금_입금전표.zip'
    return FileResponse(bundle, as_attachment=True, filename=filename)


@login_required
def export_all(request):
    output = build_workbook(job=None)
    filename = f'전체누적현황_{timezone.localdate().isoformat()}.xlsx'
    return FileResponse(output, as_attachment=True, filename=filename)


@login_required
def member_lookup(request):
    """Fast type-ahead member lookup for payment allocation and other pickers.

    Never render all 3,000+ members into a <select>. Only return the best 20
    candidates after the operator types a name, vehicle number, management
    number, phone number, or birth digits.
    """
    q = (request.GET.get('q') or '').strip()
    if not q:
        return JsonResponse({'results': []})

    digits = ''.join(ch for ch in q if ch.isdigit())
    vehicle_q = normalize_vehicle_no(q)
    query = (
        Q(name__icontains=q)
        | Q(management_no__icontains=q)
        | Q(phone__icontains=digits or q)
        | Q(birth6__icontains=digits or q)
    )
    if vehicle_q:
        query |= Q(vehicles__normalized_vehicle_no__icontains=vehicle_q)

    qs = (
        Member.objects.filter(is_active_record=True)
        .filter(query)
        .distinct()
        .order_by('name', 'management_no', 'id')[:20]
    )
    members = list(qs)
    member_ids = [member.id for member in members]
    vehicle_map = {
        row['member_id']: row['vehicle_no']
        for row in Vehicle.objects.filter(member_id__in=member_ids, is_current=True)
        .values('member_id', 'vehicle_no')
    }
    results = []
    for member in members:
        vehicle = vehicle_map.get(member.id, '')
        status = '폐업' if member.operational_status == Member.OperationalStatus.CLOSED else '현재'
        label_parts = [member.name]
        if vehicle:
            label_parts.append(vehicle)
        if member.region:
            label_parts.append(member.region)
        if member.management_no:
            label_parts.append(f'관리 {member.management_no}')
        results.append({
            'id': member.id,
            'name': member.name,
            'vehicle': vehicle,
            'region': member.region or '',
            'management_no': member.management_no or '',
            'phone': member.phone or '',
            'status': status,
            'label': ' · '.join(label_parts),
        })
    return JsonResponse({'results': results})


@login_required
def member_list(request):
    q = request.GET.get('q', '').strip()
    region = request.GET.get('region', '').strip()
    membership = request.GET.get('membership', '').strip()
    field = request.GET.get('field', 'all').strip() or 'all'
    try:
        page_size = int(request.GET.get('page_size', '50'))
    except (TypeError, ValueError):
        page_size = 50
    page_size = page_size if page_size in {50, 100} else 50

    members_qs = Member.objects.filter(
        is_active_record=True,
        operational_status=Member.OperationalStatus.ACTIVE,
    )
    if region:
        members_qs = members_qs.filter(region=region)
    if membership in {choice for choice, _ in Member.MembershipStatus.choices}:
        members_qs = members_qs.filter(membership_status=membership)
    if q:
        digits = ''.join(ch for ch in q if ch.isdigit())
        vehicle_q = normalize_vehicle_no(q)
        if field == 'name':
            members_qs = members_qs.filter(name__icontains=q)
        elif field == 'vehicle':
            members_qs = members_qs.filter(vehicles__normalized_vehicle_no__icontains=vehicle_q).distinct()
        elif field == 'phone':
            members_qs = members_qs.filter(phone__icontains=digits or q)
        elif field == 'address':
            members_qs = members_qs.filter(Q(address__icontains=q) | Q(official_address__icontains=q))
        elif field == 'management_no':
            members_qs = members_qs.filter(management_no__icontains=q)
        else:
            query = (
                Q(name__icontains=q) | Q(phone__icontains=digits or q)
                | Q(address__icontains=q) | Q(official_address__icontains=q)
                | Q(management_no__icontains=q) | Q(birth6__icontains=digits or q)
                | Q(memo__icontains=q)
            )
            if vehicle_q:
                query |= Q(vehicles__normalized_vehicle_no__icontains=vehicle_q)
            members_qs = members_qs.filter(query).distinct()

    members_qs = members_qs.order_by('region', 'management_no', 'name', 'id')
    page_obj = _attach_member_page_data(
        Paginator(members_qs, page_size).get_page(request.GET.get('page'))
    )
    counts = _member_status_counts()
    return render(request, 'core/member_list.html', {
        'page_obj': page_obj,
        'q': q,
        'region': region,
        'membership': membership,
        'field': field,
        'page_size': page_size,
        'closed_mode': False,
        'regions': _member_regions(),
        'active_count': counts['active'] or 0,
        'closed_count': counts['closed'] or 0,
    })


@login_required
def closed_member_list(request):
    q = request.GET.get('q', '').strip()
    region = request.GET.get('region', '').strip()
    field = request.GET.get('field', 'all').strip() or 'all'
    try:
        page_size = int(request.GET.get('page_size', '50'))
    except (TypeError, ValueError):
        page_size = 50
    page_size = page_size if page_size in {50, 100} else 50

    members_qs = Member.objects.filter(
        is_active_record=True,
        operational_status=Member.OperationalStatus.CLOSED,
    )
    if region:
        members_qs = members_qs.filter(region=region)
    if q:
        digits = ''.join(ch for ch in q if ch.isdigit())
        vehicle_q = normalize_vehicle_no(q)
        if field == 'name':
            members_qs = members_qs.filter(name__icontains=q)
        elif field == 'vehicle':
            members_qs = members_qs.filter(vehicles__normalized_vehicle_no__icontains=vehicle_q).distinct()
        elif field == 'phone':
            members_qs = members_qs.filter(phone__icontains=digits or q)
        elif field == 'address':
            members_qs = members_qs.filter(Q(address__icontains=q) | Q(official_address__icontains=q))
        elif field == 'management_no':
            members_qs = members_qs.filter(management_no__icontains=q)
        else:
            query = (
                Q(name__icontains=q) | Q(phone__icontains=digits or q)
                | Q(address__icontains=q) | Q(official_address__icontains=q)
                | Q(management_no__icontains=q) | Q(birth6__icontains=digits or q)
                | Q(memo__icontains=q)
            )
            if vehicle_q:
                query |= Q(vehicles__normalized_vehicle_no__icontains=vehicle_q)
            members_qs = members_qs.filter(query).distinct()

    members_qs = members_qs.order_by('-closed_on', 'region', 'management_no', 'name', 'id')
    page_obj = _attach_member_page_data(
        Paginator(members_qs, page_size).get_page(request.GET.get('page'))
    )
    counts = _member_status_counts()
    return render(request, 'core/member_list.html', {
        'page_obj': page_obj,
        'q': q,
        'region': region,
        'membership': '',
        'field': field,
        'page_size': page_size,
        'closed_mode': True,
        'regions': _member_regions(),
        'active_count': counts['active'] or 0,
        'closed_count': counts['closed'] or 0,
    })


@login_required
def member_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    mode = request.GET.get('mode', 'active')
    closed = mode == 'closed'
    status = Member.OperationalStatus.CLOSED if closed else Member.OperationalStatus.ACTIVE
    rows = _member_queryset_with_financials(
        Member.objects.filter(is_active_record=True, operational_status=status)
    ).order_by('region', 'current_vehicle_no', 'name', 'id')

    wb = Workbook()
    ws = wb.active
    ws.title = '폐업명단' if closed else '현재명단'
    if closed:
        headers = ['관리번호', '지역', '차량번호', '성명', '생년월일', '핸드폰', '폐업일', '주소', '잔액']
    else:
        headers = ['관리번호', '지역', '차량번호', '성명', '생년월일', '핸드폰', '가입상태', '가입일', '자격증명 발급일', '주소', '잔액']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='28314A')
        cell.alignment = Alignment(horizontal='center')

    for member in rows.iterator(chunk_size=500):
        common = [
            member.management_no or '', member.region or '', member.current_vehicle_no or '', member.name,
            member.birth6 or '', member.phone or '',
        ]
        if closed:
            values = common + [member.closed_on, member.address or '', float((member.outstanding_amount or 0) - (member.prepayment_amount or 0))]
        else:
            values = common + [member.get_membership_status_display(), member.membership_started_on, member.certificate_issued_on, member.address or '', float((member.outstanding_amount or 0) - (member.prepayment_amount or 0))]
        ws.append(values)

    widths = [16, 12, 18, 12, 12, 16, 12, 13, 16, 42, 14]
    for index, width in enumerate(widths[:len(headers)], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = '폐업명단.xlsx' if closed else '현재명단.xlsx'
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    return response


@login_required
def member_detail(request, pk):
    member = get_object_or_404(Member, pk=pk)
    current_vehicle = Vehicle.objects.filter(member=member, is_current=True).order_by('-start_date', '-id').first()
    charges = list(
        member.charges.filter(
            Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True)
        ).annotate(
            settled_amount_fast=Coalesce(
                Sum('settlements__amount', filter=Q(settlements__is_active=True)),
                MONEY_ZERO_VALUE,
            )
        ).order_by('-charge_date', '-id')[:100]
    )
    for charge in charges:
        charge.balance_fast = max(Decimal('0'), charge.amount - charge.settled_amount_fast)

    charge_total = member.charges.filter(
        status=Charge.Status.POSTED,
    ).filter(
        Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True)
    ).aggregate(v=Sum('amount'))['v'] or Decimal('0')
    settlement_total = ChargeSettlement.objects.filter(
        charge__member=member, charge__status=Charge.Status.POSTED, is_active=True,
    ).filter(
        Q(charge__monthly_job__isnull=True) | Q(charge__monthly_job__is_current=True)
    ).aggregate(v=Sum('amount'))['v'] or Decimal('0')
    outstanding_amount = max(Decimal('0'), charge_total - settlement_total)
    prepayment_total = member.prepayments.filter(balance__gt=0).aggregate(v=Sum('balance'))['v'] or Decimal('0')
    net_balance = outstanding_amount - prepayment_total

    payments = member.payment_allocation_lines.filter(
        status=PaymentAllocationLine.Status.ACTIVE
    ).select_related('payment').order_by('-payment__payment_date')[:100]
    return render(request, 'core/member_detail.html', {
        'member': member,
        'current_vehicle': current_vehicle,
        'outstanding_amount': outstanding_amount,
        'prepayment_total': prepayment_total,
        'net_balance': net_balance,
        'charges': charges,
        'payments': payments,
        'prepayments': member.prepayments.all(),
        'refunds': member.refunds.all(),
        'messages_history': member.message_recipients.select_related('batch').order_by('-created_at')[:50],
        'notices': member.legal_notices.all()[:50],
        'audits': AuditLog.objects.filter(model_name__icontains='member', object_id=str(member.id))[:50],
        'outgoing_links': member.outgoing_links.select_related('new_member').all(),
        'incoming_links': member.incoming_links.select_related('old_member').all(),
    })


@login_required
def member_payment_history(request, pk):
    member = get_object_or_404(Member, pk=pk)
    year = 2026
    start_month = 1
    end_month = 7
    legacy = list(
        HistoricalPaymentRecord.objects.filter(
            member=member, year=year, month__gte=start_month, month__lte=end_month,
        ).order_by('month', 'payment_date', 'id')
    )
    monthly = {month: Decimal('0') for month in range(start_month, end_month + 1)}
    for item in legacy:
        monthly[item.month] += item.amount
    month_rows = [{'month': m, 'amount': monthly[m]} for m in range(start_month, end_month + 1)]

    # Live-program payments are shown separately so old source facts never change the ledger.
    live_lines = list(
        member.payment_allocation_lines.filter(
            status=PaymentAllocationLine.Status.ACTIVE,
            payment__is_effective=True,
            payment__payment_date__year=year,
            payment__payment_date__month__gte=start_month,
            payment__payment_date__month__lte=end_month,
        ).select_related('payment').order_by('payment__payment_date', 'id')
    )
    legacy_total = sum((item.amount for item in legacy), Decimal('0'))
    return render(request, 'core/partials/member_payment_history_modal.html', {
        'member': member,
        'year': year,
        'start_month': start_month,
        'end_month': end_month,
        'legacy': legacy,
        'legacy_total': legacy_total,
        'month_rows': month_rows,
        'live_lines': live_lines,
    })


@login_required
def member_create(request):
    initial = {
        'name': request.GET.get('name', ''),
        'vehicle_no': request.GET.get('vehicle_no', ''),
        'region': request.GET.get('region', ''),
    }
    form = QuickMemberForm(request.POST or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        member = form.save()
        _invalidate_member_list_cache()
        log_action(action='member_created_quick', instance=member, actor=_actor(request))
        messages.success(request, '현재명단에 추가했습니다. 최초 부과일은 가입일·자격증명 발급일 한 달 뒤 같은 날짜로 계산됩니다.')
        return redirect('core:member_detail', pk=member.pk)
    return render(request, 'core/form.html', {
        'form': form,
        'title': '신규 명단 간편추가',
        'subtitle': '지역·차량번호·성명만 우선 입력하고 나머지 정보는 나중에 보완할 수 있습니다.',
    })


@login_required
def member_edit(request, pk):
    member = get_object_or_404(Member, pk=pk)
    before = json_safe_model(member)
    form = MemberForm(request.POST or None, instance=member)
    modal = _wants_modal(request)
    if request.method == 'POST' and form.is_valid():
        member = form.save()
        _invalidate_member_list_cache()
        log_action(action='member_updated', instance=member, before=before, actor=_actor(request))
        if modal:
            return JsonResponse({'ok': True, 'message': '회원정보를 수정했습니다.'})
        messages.success(request, '회원정보를 수정했습니다.')
        return redirect('core:member_detail', pk=member.pk)
    context = {'form': form, 'member': member, 'title': f'{member.name} 회원정보 수정'}
    if modal:
        return _modal_form_response(
            request, 'core/partials/member_edit_modal.html', context,
            status=422 if request.method == 'POST' else 200,
        )
    return render(request, 'core/member_form.html', context)


@login_required
def member_close(request, pk):
    member = get_object_or_404(Member, pk=pk)
    modal = _wants_modal(request)
    action_mode = (request.POST.get('action') or request.GET.get('action') or '').strip()
    if modal and request.method == 'GET' and action_mode not in {'close', 'move'}:
        return _modal_form_response(
            request, 'core/partials/member_action_select_modal.html', {'member': member}, status=200,
        )

    if action_mode not in {'close', 'move'}:
        action_mode = 'close'
    initial_reason = '타 지역 이관' if action_mode == 'move' else ''
    form = CloseMemberForm(
        request.POST or None,
        initial={'closure_date': timezone.localdate(), 'reason': initial_reason},
    )
    if request.method == 'POST' and form.is_valid():
        try:
            reason = form.cleaned_data['reason'].strip()
            if action_mode == 'move' and not reason:
                reason = '타 지역 이관'
            close_member(member, actor=_actor(request), **{
                'closure_date': form.cleaned_data['closure_date'],
                'reason': reason,
                'memo': form.cleaned_data['memo'],
            })
            _invalidate_member_list_cache()
            message = '이관 처리했습니다.' if action_mode == 'move' else '폐업 처리했습니다.'
            if modal:
                return JsonResponse({'ok': True, 'message': message})
            messages.success(request, message)
            return redirect('core:member_detail', pk=member.pk)
        except Exception as exc:
            if modal:
                form.add_error(None, str(exc))
            else:
                messages.error(request, str(exc))
    context = {
        'form': form, 'member': member,
        'title': f'{member.name} {"이관" if action_mode == "move" else "폐업"} 처리',
        'action_mode': action_mode,
    }
    if modal:
        return _modal_form_response(
            request, 'core/partials/member_close_modal.html', context,
            status=422 if request.method == 'POST' else 200,
        )
    return render(request, 'core/form.html', {'form': form, 'title': context['title']})


@login_required
@transaction.atomic
def member_manual_payment(request, pk):
    member = get_object_or_404(Member, pk=pk)
    form = ManualPaymentForm(request.POST or None, member=member)
    modal = _wants_modal(request)
    if request.method == 'POST' and form.is_valid():
        request_key = form.cleaned_data.get('request_key') or ''
        session_key = f'manual-payment:{request_key}' if request_key else ''
        if session_key and request.session.get(session_key):
            if modal:
                return JsonResponse({'ok': True, 'message': '이미 반영된 입금입니다.'})
            messages.info(request, '이미 반영된 입금입니다.')
            return redirect('core:member_detail', pk=member.pk)

        payment_day = form.cleaned_data['payment_date']
        payment_dt = timezone.make_aware(
            datetime.combine(payment_day, time(hour=12)),
            timezone.get_current_timezone(),
        )
        job = get_or_create_current_job(payment_day)
        payer_name = (form.cleaned_data.get('payer_name') or member.name).strip()
        user_memo = (form.cleaned_data.get('memo') or '').strip()
        memo_parts = [f'입금자 {payer_name}']
        if user_memo:
            memo_parts.append(user_memo)
        payment = Payment.objects.create(
            source_type=Payment.SourceType.MANUAL,
            payment_date=payment_dt,
            amount=form.cleaned_data['amount'],
            monthly_job=job,
            memo=' · '.join(memo_parts),
            status=Payment.Status.OPEN,
            is_effective=True,
        )
        replace_payment_allocations(
            payment,
            [{
                'member': member,
                'account_type': form.cleaned_data['account_type'],
                'amount': form.cleaned_data['amount'],
                'memo': '회원명단에서 수기입금',
            }],
            reason='회원명단 수기입금',
            actor=_actor(request),
        )
        log_action(action='manual_payment_created', instance=payment, actor=_actor(request))
        if session_key:
            request.session[session_key] = payment.id
        if modal:
            return JsonResponse({'ok': True, 'message': f'{form.cleaned_data["amount"]:,.0f}원 입금을 반영했습니다.'})
        messages.success(request, '수기입금을 반영했습니다.')
        return redirect('core:member_detail', pk=member.pk)

    context = {'form': form, 'member': member, 'title': f'{member.name} 수기입금'}
    if modal:
        return _modal_form_response(
            request, 'core/partials/member_payment_modal.html', context,
            status=422 if request.method == 'POST' else 200,
        )
    return render(request, 'core/form.html', {'form': form, 'title': f'{member.name} 수기입금'})


@login_required
def member_reopen(request, pk):
    member = get_object_or_404(Member, pk=pk)
    form = ReopenMemberForm(request.POST or None, initial={'re_registered_on': timezone.localdate()})
    if request.method == 'POST' and form.is_valid():
        reopen_member(member, actor=_actor(request), **form.cleaned_data)
        _invalidate_member_list_cache()
        messages.success(request, '재등록 처리했습니다.')
        return redirect('core:member_detail', pk=member.pk)
    return render(request, 'core/form.html', {'form': form, 'title': f'{member.name} 재등록'})


@login_required
def member_join(request, pk):
    member = get_object_or_404(Member, pk=pk)
    form = JoinAssociationForm(request.POST or None, initial={'join_date': timezone.localdate()})
    if request.method == 'POST' and form.is_valid():
        try:
            completed = join_association(member, actor=_actor(request), **form.cleaned_data)
            if completed:
                messages.success(request, '협회가입 처리했습니다.')
            else:
                messages.warning(request, '관리비 미수금이 남아 가입대기로 처리했습니다. 전액 납부 후 다시 가입 완료하세요.')
            return redirect('core:member_detail', pk=member.pk)
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, 'core/form.html', {'form': form, 'title': f'{member.name} 협회가입'})


@login_required
def member_leave(request, pk):
    member = get_object_or_404(Member, pk=pk)
    form = LeaveAssociationForm(request.POST or None, initial={'leave_date': timezone.localdate()})
    if request.method == 'POST' and form.is_valid():
        try:
            leave_association(member, actor=_actor(request), **form.cleaned_data)
            messages.success(request, '협회탈퇴 처리했습니다.')
            return redirect('core:member_detail', pk=member.pk)
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, 'core/form.html', {'form': form, 'title': f'{member.name} 협회탈퇴'})


@login_required
def member_transfer(request, pk):
    member = get_object_or_404(Member, pk=pk)
    current_vehicle = member.current_vehicle
    today = timezone.localdate()
    initial = {
        'received_date': today,
        'approval_date': today,
        'effective_date': today,
        'new_vehicle_no': current_vehicle.vehicle_no if current_vehicle else '',
        'new_region': member.region,
    }
    form = TransferMemberForm(request.POST or None, initial=initial)
    modal = _wants_modal(request)
    if request.method == 'POST' and form.is_valid():
        try:
            memo_bits = []
            if form.cleaned_data.get('received_date'):
                memo_bits.append(f"접수일 {form.cleaned_data['received_date']:%Y-%m-%d}")
            if form.cleaned_data.get('approval_date'):
                memo_bits.append(f"인가일 {form.cleaned_data['approval_date']:%Y-%m-%d}")
            if form.cleaned_data.get('memo'):
                memo_bits.append(form.cleaned_data['memo'].strip())
            new_member, link = transfer_member(
                member,
                transfer_type=MemberLink.LinkType.GENERAL_TRANSFER,
                effective_date=form.cleaned_data['effective_date'],
                new_name=form.cleaned_data['new_name'],
                new_birth6=form.cleaned_data.get('new_birth6', ''),
                new_phone=form.cleaned_data.get('new_phone', ''),
                new_address=form.cleaned_data.get('new_address', ''),
                new_official_address='',
                new_vehicle_no=form.cleaned_data.get('new_vehicle_no', ''),
                new_region=form.cleaned_data.get('new_region', ''),
                memo=' · '.join(memo_bits),
                actor=_actor(request),
            )
            if form.cleaned_data.get('approval_date'):
                new_member.certificate_issued_on = form.cleaned_data['approval_date']
                new_member.save(update_fields=['certificate_issued_on', 'updated_at'])
            _invalidate_member_list_cache()
            message = '도내 양도양수를 처리했습니다. 기존 미수금은 이전 명의자에게 유지됩니다.'
            if modal:
                return JsonResponse({'ok': True, 'message': message})
            messages.success(request, message)
            return redirect('core:member_detail', pk=new_member.pk)
        except Exception as exc:
            if modal:
                form.add_error(None, str(exc))
            else:
                messages.error(request, str(exc))
    context = {'form': form, 'member': member, 'title': f'{member.name} 도내 양도양수'}
    if modal:
        return _modal_form_response(
            request, 'core/partials/member_transfer_modal.html', context,
            status=422 if request.method == 'POST' else 200,
        )
    return render(request, 'core/form.html', {'form': form, 'title': context['title']})


@login_required
def bank_transaction_list(request):
    job_id = request.GET.get('job')
    status = request.GET.get('status', '')
    q = request.GET.get('q', '').strip()
    txs = BankTransaction.objects.select_related('job', 'uploaded_file', 'payment').filter(job__is_current=True)
    if job_id:
        txs = txs.filter(job_id=job_id)
    if status == 'review':
        txs = txs.filter(status__in=[BankTransaction.Status.UNMATCHED, BankTransaction.Status.REVIEW, BankTransaction.Status.DUPLICATE])
    elif status:
        txs = txs.filter(status=status)
    if q:
        txs = txs.filter(Q(payer_text__icontains=q) | Q(bank_account_label__icontains=q))
    return render(request, 'core/bank_list.html', {
        'transactions': txs.order_by('-transaction_at', '-id')[:2000],
        'jobs': MonthlyJob.objects.filter(is_current=True), 'status_choices': BankTransaction.Status.choices,
        'job_id': job_id, 'status': status, 'q': q,
    })


@login_required
def card_transaction_list(request):
    job_id = request.GET.get('job')
    status = request.GET.get('status', '')
    q = request.GET.get('q', '').strip()
    txs = CardTransaction.objects.select_related('job', 'uploaded_file', 'payment').filter(job__is_current=True)
    if job_id:
        txs = txs.filter(job_id=job_id)
    if status:
        txs = txs.filter(status=status)
    if q:
        txs = txs.filter(
            Q(member_name__icontains=q) | Q(vehicle_no__icontains=q) | Q(txn_key__icontains=q)
        )
    return render(request, 'core/card_list.html', {
        'transactions': txs.order_by('-transaction_at', '-id')[:2000],
        'jobs': MonthlyJob.objects.filter(is_current=True),
        'status_choices': CardTransaction.Status.choices,
        'job_id': job_id, 'status': status, 'q': q,
        'altolan_form': SimpleExcelUploadForm(),
        'cider_form': SimpleExcelUploadForm(),
    })


@login_required
@transaction.atomic
def card_upload(request, provider):
    if request.method != 'POST':
        raise Http404
    provider_to_slot = {
        'altolan': UploadedFile.SlotType.ALTOLAN,
        'cider': UploadedFile.SlotType.CIDER,
    }
    slot_type = provider_to_slot.get(provider)
    if not slot_type:
        raise Http404
    form = SimpleExcelUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, '엑셀 파일을 선택하세요.')
        return redirect('core:card_transaction_list')
    try:
        job = get_or_create_current_job(timezone.localdate())
        uploaded, result = _save_and_process_excel(
            job=job, slot_type=slot_type, file_obj=form.cleaned_data['file'],
        )
        if result.get('already_processed'):
            messages.info(request, '이미 올린 같은 파일이라 다시 반영하지 않았습니다.')
        else:
            auto_match_card_job(job)
            messages.success(
                request,
                f'{uploaded.get_slot_type_display()} 반영 완료: '
                f"신규 {result.get('created_transactions', 0)}건 · "
                f"중복의심 {result.get('duplicates', 0)}건",
            )
    except Exception as exc:
        messages.error(request, f'카드결제 반영 실패: {exc}')
    return redirect('core:card_transaction_list')


@login_required
def payment_allocate(request, pk):
    payment = get_object_or_404(Payment.objects.select_related('bank_transaction', 'card_transaction'), pk=pk)
    existing = payment.allocation_lines.filter(status=PaymentAllocationLine.Status.ACTIVE)
    initial = [{
        'member': line.member_id, 'account_type': line.account_type,
        'amount': line.amount, 'memo': line.memo,
    } for line in existing]
    formset = AllocationFormSet(request.POST or None, initial=initial)
    if request.method == 'POST' and formset.is_valid():
        rows = []
        for form in formset:
            if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                continue
            rows.append({
                'member': form.cleaned_data['member'],
                'account_type': form.cleaned_data['account_type'],
                'amount': form.cleaned_data['amount'],
                'memo': form.cleaned_data.get('memo', ''),
            })
        try:
            replace_payment_allocations(payment, rows, reason='화면 수동배정', actor=_actor(request))
            if request.POST.get('save_payer_alias') and payment.bank_transaction_id and len(rows) == 1:
                payer = payment.bank_transaction.payer_text.strip()
                if payer:
                    PayerAlias.objects.update_or_create(
                        member=rows[0]['member'],
                        normalized_alias=normalize_text(payer),
                        bank_account_label=payment.bank_transaction.bank_account_label,
                        defaults={
                            'alias': payer,
                            'auto_apply': True,
                            'memo': '입금 배정 화면에서 저장',
                            'actor': _actor(request),
                        },
                    )
            messages.success(request, '입금 배정을 저장했습니다.')
            if payment.bank_transaction_id:
                return redirect('core:bank_transaction_list')
            return redirect('core:dashboard')
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, 'core/payment_allocate.html', {'payment': payment, 'formset': formset})


@login_required
def charge_list(request):
    account = request.GET.get('account', '')
    region = request.GET.get('region', '')
    q = request.GET.get('q', '').strip()
    min_amount = request.GET.get('min_amount', '').strip()
    max_amount = request.GET.get('max_amount', '').strip()
    balance_state = request.GET.get('balance_state', 'outstanding')

    charges = (
        Charge.objects.filter(Q(monthly_job__isnull=True) | Q(monthly_job__is_current=True))
        .filter(status=Charge.Status.POSTED)
        .select_related('member', 'monthly_job')
        .annotate(
            settled_total=Coalesce(
                Sum('settlements__amount', filter=Q(settlements__is_active=True)),
                MONEY_ZERO_VALUE,
            ),
        )
        .annotate(
            balance_amount=ExpressionWrapper(
                F('amount') - F('settled_total'), output_field=MONEY_FIELD,
            ),
        )
    )
    if account:
        charges = charges.filter(account_type=account)
    if region:
        charges = charges.filter(member__region=region)
    if q:
        charges = charges.filter(
            Q(member__name__icontains=q) | Q(member__vehicles__vehicle_no__icontains=q)
        ).distinct()
    if balance_state == 'outstanding':
        charges = charges.filter(balance_amount__gt=0)
    elif balance_state == 'paid':
        charges = charges.filter(balance_amount__lte=0)
    try:
        if min_amount:
            charges = charges.filter(balance_amount__gte=Decimal(min_amount.replace(',', '')))
        if max_amount:
            charges = charges.filter(balance_amount__lte=Decimal(max_amount.replace(',', '')))
    except Exception:
        messages.warning(request, '금액 필터는 숫자로 입력하세요.')

    regions = list(
        Member.objects.filter(is_active_record=True).exclude(region='')
        .values_list('region', flat=True).distinct().order_by('region')
    )
    paginator = Paginator(charges.order_by('-balance_amount', 'member__region', 'member__name'), 100)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'core/charge_list.html', {
        'charges': page_obj, 'page_obj': page_obj,
        'account_choices': [
            (AccountType.MEMBERSHIP_FEE, '협회비'),
            (AccountType.MANAGEMENT_FEE, '관리비'),
        ],
        'account': account, 'region': region, 'regions': regions, 'q': q,
        'min_amount': min_amount, 'max_amount': max_amount, 'balance_state': balance_state,
    })


@login_required
def refund_list(request):
    status = request.GET.get('status', '')
    refunds = Refund.objects.select_related('member')
    if status:
        refunds = refunds.filter(status=status)
    return render(request, 'core/refund_list.html', {'refunds': refunds, 'status_choices': Refund.Status.choices, 'status': status})


@login_required
def refund_create(request, member_pk):
    member = get_object_or_404(Member, pk=member_pk)
    form = RefundPendingForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        refund = form.save(commit=False)
        refund.member = member
        refund.status = Refund.Status.PENDING
        refund.save()
        log_action(action='refund_pending_created', instance=refund, actor=_actor(request))
        messages.success(request, '환불대기를 등록했습니다.')
        return redirect('core:member_detail', pk=member.pk)
    return render(request, 'core/form.html', {'form': form, 'title': f'{member.name} 환불대기 등록'})


@login_required
def refund_complete(request, pk):
    refund = get_object_or_404(Refund, pk=pk)
    form = RefundForm(request.POST or None, instance=refund)
    if request.method == 'POST' and form.is_valid():
        updated = form.save(commit=False)
        updated.status = Refund.Status.PENDING
        updated.save()
        try:
            complete_refund(updated, actor=_actor(request))
            if form.cleaned_data['send_completion_sms']:
                batch = create_refund_batch(
                    [updated.id], message_type=MessageTemplate.TemplateType.REFUND_COMPLETE,
                    immediate=True, actor=_actor(request),
                )
                send_batch(batch, actor=_actor(request))
            messages.success(request, '환불 완료 처리했습니다.')
            return redirect('core:member_detail', pk=refund.member_id)
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, 'core/form.html', {'form': form, 'title': f'{refund.member.name} 환불 완료'})


@login_required
def arrears_compose(request):
    q = request.GET.get('q', '').strip()
    operational = request.GET.get('operational', '')
    contact = request.GET.get('contact', '')
    collection = request.GET.get('collection', '')
    region = request.GET.get('region', '').strip()
    account = request.GET.get('account', '').strip()
    membership = request.GET.get('membership', '').strip()
    amount_band = request.GET.get('amount_band', '').strip()
    min_amount = request.GET.get('min_amount', '').strip()
    max_amount = request.GET.get('max_amount', '').strip()

    members_qs = Member.objects.filter(is_active_record=True).order_by('name')
    if q:
        members_qs = members_qs.filter(
            Q(name__icontains=q) | Q(phone__icontains=q) | Q(vehicles__vehicle_no__icontains=q)
        ).distinct()
    if operational:
        members_qs = members_qs.filter(operational_status=operational)
    if region:
        members_qs = members_qs.filter(region=region)
    if account in {AccountType.MEMBERSHIP_FEE, AccountType.MANAGEMENT_FEE}:
        members_qs = members_qs.filter(receivable_account_type=account)
    if membership in {choice for choice, _ in Member.MembershipStatus.choices}:
        members_qs = members_qs.filter(membership_status=membership)
    if collection:
        members_qs = members_qs.filter(collection_status=collection)
    if contact == 'can_sms':
        members_qs = members_qs.exclude(phone='').filter(phone_needs_check=False, sms_opt_out=False)
    elif contact == 'missing_phone':
        members_qs = members_qs.filter(Q(phone='') | Q(phone_needs_check=True))
    elif contact == 'opt_out':
        members_qs = members_qs.filter(sms_opt_out=True)
    elif contact == 'address':
        members_qs = members_qs.filter(address_needs_check=True)
    elif contact == 'not_sent':
        members_qs = members_qs.exclude(message_recipients__status__in=[MessageRecipient.Status.ACCEPTED, MessageRecipient.Status.SENT]).distinct()

    financial_qs = _member_queryset_with_financials(members_qs).filter(outstanding_amount__gt=0)
    band_map = {
        'under_50k': (None, Decimal('49999')),
        '50k_100k': (Decimal('50000'), Decimal('99999')),
        '100k_200k': (Decimal('100000'), Decimal('199999')),
        '200k_500k': (Decimal('200000'), Decimal('499999')),
        '500k_plus': (Decimal('500000'), None),
    }
    band_min, band_max = band_map.get(amount_band, (None, None))
    try:
        min_value = Decimal(min_amount.replace(',', '')) if min_amount else band_min
        max_value = Decimal(max_amount.replace(',', '')) if max_amount else band_max
        if min_value is not None:
            financial_qs = financial_qs.filter(outstanding_amount__gte=min_value)
        if max_value is not None:
            financial_qs = financial_qs.filter(outstanding_amount__lte=max_value)
    except Exception:
        messages.warning(request, '미수금액 필터는 숫자로 입력하세요.')

    candidates = list(financial_qs.order_by('-outstanding_amount', 'name', 'id')[:1500])
    form = ArrearsComposeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        selected = request.POST.getlist('member_ids')
        if not selected:
            messages.error(request, '대상자를 선택하세요.')
        else:
            scheduled_at = form.cleaned_data['scheduled_at']
            batch = create_arrears_batch(
                selected,
                due_date=form.cleaned_data['due_date'],
                scheduled_at=scheduled_at,
                immediate=not bool(scheduled_at),
                actor=_actor(request),
            )
            return redirect('core:message_batch_detail', pk=batch.pk)
    return render(request, 'core/arrears_compose.html', {
        'form': form, 'candidates': candidates, 'q': q,
        'operational': operational, 'contact': contact, 'collection': collection,
        'region': region, 'regions': _member_regions(), 'account': account,
        'membership': membership, 'amount_band': amount_band,
        'min_amount': min_amount, 'max_amount': max_amount,
        'operational_choices': Member.OperationalStatus.choices,
        'collection_choices': Member.CollectionStatus.choices,
        'membership_choices': Member.MembershipStatus.choices,
        'account_choices': [
            (AccountType.MEMBERSHIP_FEE, '협회비'),
            (AccountType.MANAGEMENT_FEE, '관리비'),
        ],
    })


@login_required
def message_list(request):
    client = BalsongClient()
    if request.method == 'POST' and request.POST.get('action') == 'check_provider':
        try:
            result = client.callback_list()
            callbacks = [str(item.get('CallBack_No') or '') for item in result.get('List') or []]
            callback_digits = ''.join(ch for ch in client.callback if ch.isdigit())
            registered_digits = {
                ''.join(ch for ch in number if ch.isdigit()) for number in callbacks
            }
            if callback_digits and callback_digits not in registered_digits:
                messages.warning(
                    request,
                    f'아이디·비밀번호 연결은 성공했지만 발신번호 {client.callback}가 등록목록에 없습니다.',
                )
            else:
                mode = '시험모드' if client.dry_run else '실전송 모드'
                messages.success(
                    request,
                    f'발송닷컴 연결 정상 · {mode} · 잔액 {int(result.get("Cash") or 0):,}원 · 등록 발신번호 {len(callbacks)}개',
                )
        except Exception as exc:
            messages.error(request, f'발송닷컴 연결 실패: {exc}')
        return redirect('core:message_list')

    return render(request, 'core/message_list.html', {
        'batches': MessageBatch.objects.all()[:500],
        'balsong_dry_run': client.dry_run,
        'balsong_callback': client.callback,
    })


@login_required
def message_batch_detail(request, pk):
    batch = get_object_or_404(MessageBatch, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'update_selection':
                if batch.provider_job_no or batch.status not in {
                    MessageBatch.Status.DRAFT,
                    MessageBatch.Status.SCHEDULED,
                }:
                    raise ValueError('발송닷컴 접수 후에는 대상자를 변경할 수 없습니다.')
                selected = set(request.POST.getlist('recipient_ids'))
                for recipient in batch.recipients.all():
                    if recipient.exclusion_reason and recipient.exclusion_reason != '이번 발송 제외':
                        continue
                    if str(recipient.id) in selected:
                        recipient.status = MessageRecipient.Status.PENDING
                        recipient.exclusion_reason = ''
                    else:
                        recipient.status = MessageRecipient.Status.EXCLUDED
                        recipient.exclusion_reason = '이번 발송 제외'
                    recipient.save(update_fields=['status', 'exclusion_reason', 'updated_at'])
                messages.success(request, '이번 발송 대상을 수정했습니다.')

            elif action == 'confirm':
                batch = send_batch(batch, actor=_actor(request))
                if batch.status == MessageBatch.Status.DRY_RUN:
                    messages.warning(request, '시험모드 처리만 했습니다. 실제 문자는 발송되지 않았습니다.')
                elif batch.status == MessageBatch.Status.SCHEDULED:
                    messages.success(
                        request,
                        f'발송닷컴 예약 접수 완료 · 접수번호 {batch.provider_job_no}',
                    )
                elif batch.status == MessageBatch.Status.ACCEPTED:
                    messages.success(
                        request,
                        f'발송닷컴 접수 완료 · 접수번호 {batch.provider_job_no}. 전송결과는 결과 확인 버튼으로 갱신합니다.',
                    )
                elif batch.status == MessageBatch.Status.FAILED:
                    messages.error(
                        request,
                        (batch.provider_response or {}).get('message', '발송 요청이 실패했습니다.'),
                    )

            elif action == 'sync_results':
                batch = sync_batch_results(batch, actor=_actor(request))
                messages.success(request, f'발송결과를 확인했습니다. 현재 상태: {batch.get_status_display()}')

            elif action == 'cancel':
                cancel_batch(batch, actor=_actor(request))
                messages.success(request, '예약발송을 취소했습니다.')

            elif action == 'retry_failed':
                if batch.provider_job_no and batch.provider_job_no != 'DRY-RUN':
                    sync_batch_results(batch, actor=_actor(request))
                retry = retry_failed_batch(batch, actor=_actor(request))
                return redirect('core:message_batch_detail', pk=retry.pk)
        except Exception as exc:
            messages.error(request, str(exc))
        return redirect('core:message_batch_detail', pk=batch.pk)

    recipients = batch.recipients.select_related('member')
    return render(request, 'core/message_batch_detail.html', {
        'batch': batch,
        'recipients': recipients,
        'can_edit_selection': not batch.provider_job_no and batch.status in {
            MessageBatch.Status.DRAFT,
            MessageBatch.Status.SCHEDULED,
        },
        'can_sync': bool(batch.provider_job_no and batch.provider_job_no != 'DRY-RUN'),
        'can_cancel': batch.status == MessageBatch.Status.SCHEDULED,
        'has_failed': batch.recipients.filter(status=MessageRecipient.Status.FAILED).exists(),
        'preview_recipient': batch.recipients.exclude(status=MessageRecipient.Status.EXCLUDED).first(),
    })


@login_required
def message_batch_edit(request, pk):
    batch = get_object_or_404(MessageBatch, pk=pk)
    if batch.status not in {MessageBatch.Status.DRAFT, MessageBatch.Status.SCHEDULED}:
        messages.error(request, '발송 완료 후에는 예약정보를 수정할 수 없습니다.')
        return redirect('core:message_batch_detail', pk=batch.pk)

    form = MessageScheduleEditForm(request.POST or None, instance=batch)
    if request.method == 'POST' and form.is_valid():
        try:
            scheduled_at = form.cleaned_data['scheduled_at']
            due_date = form.cleaned_data['due_date']
            if batch.provider_job_no:
                update_batch_schedule(
                    batch,
                    scheduled_at=scheduled_at,
                    due_date=due_date,
                    actor=_actor(request),
                )
            else:
                batch.scheduled_at = scheduled_at
                batch.due_date = due_date
                batch.status = (
                    MessageBatch.Status.SCHEDULED
                    if scheduled_at and scheduled_at > timezone.now()
                    else MessageBatch.Status.DRAFT
                )
                batch.save(update_fields=['scheduled_at', 'due_date', 'status', 'updated_at'])
                log_action(action='message_schedule_updated', instance=batch, actor=_actor(request))
            messages.success(request, '예약정보를 수정했습니다.')
            return redirect('core:message_batch_detail', pk=batch.pk)
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, 'core/form.html', {'form': form, 'title': '문자 예약정보 수정'})


@login_required
def refund_message_compose(request):
    refunds = Refund.objects.select_related('member').exclude(status=Refund.Status.CANCELLED).order_by('-created_at')
    form = RefundMessageForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        selected = request.POST.getlist('refund_ids')
        if not selected:
            messages.error(request, '환불 건을 선택하세요.')
        else:
            batch = create_refund_batch(
                selected,
                message_type=form.cleaned_data['message_type'],
                scheduled_at=form.cleaned_data['scheduled_at'],
                immediate=True,
                actor=_actor(request),
            )
            return redirect('core:message_batch_detail', pk=batch.pk)
    return render(request, 'core/refund_message_compose.html', {'form': form, 'refunds': refunds})


@login_required
def legal_notice_list(request):
    return render(request, 'core/legal_notice_list.html', {'notices': LegalNotice.objects.select_related('member')[:1000]})


@login_required
def legal_notice_create(request, member_pk):
    member = get_object_or_404(Member, pk=member_pk)
    if member.address_needs_check:
        messages.error(request, '주소 확인 필요 상태이므로 새 내용증명을 등록할 수 없습니다.')
        return redirect('core:member_detail', pk=member.pk)
    form = LegalNoticeForm(request.POST or None, initial={'sent_date': timezone.localdate()})
    if request.method == 'POST' and form.is_valid():
        notice = form.save(commit=False)
        notice.member = member
        notice.actor = _actor(request)
        if notice.address_type == LegalNotice.AddressType.BASIC:
            notice.address_snapshot = member.address
        elif notice.address_type == LegalNotice.AddressType.OFFICIAL:
            notice.address_snapshot = member.official_address
        else:
            notice.address_snapshot = member.address
            notice.second_address_snapshot = '' if member.address == member.official_address else member.official_address
        notice.save()
        log_action(action='legal_notice_created', instance=notice, actor=_actor(request))
        return redirect('core:member_detail', pk=member.pk)
    return render(request, 'core/form.html', {'form': form, 'title': f'{member.name} 내용증명 등록'})


@login_required
def legal_notice_edit(request, pk):
    notice = get_object_or_404(LegalNotice, pk=pk)
    before_status = notice.delivery_status
    form = LegalNoticeForm(request.POST or None, instance=notice)
    if request.method == 'POST' and form.is_valid():
        notice = form.save()
        failed_statuses = {
            LegalNotice.DeliveryStatus.RETURNED,
            LegalNotice.DeliveryStatus.UNKNOWN_RECIPIENT,
            LegalNotice.DeliveryStatus.UNKNOWN_ADDRESS,
            LegalNotice.DeliveryStatus.ABSENT,
        }
        if notice.delivery_status in failed_statuses:
            notice.member.address_needs_check = True
            notice.member.save(update_fields=['address_needs_check', 'updated_at'])
        log_action(action='legal_notice_updated', instance=notice, before={'delivery_status': before_status}, actor=_actor(request))
        return redirect('core:member_detail', pk=notice.member_id)
    return render(request, 'core/form.html', {'form': form, 'title': '내용증명 배송결과 수정'})


@login_required
def address_check_clear(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        member.address_needs_check = False
        member.save(update_fields=['address_needs_check', 'updated_at'])
        log_action(action='address_check_cleared', instance=member, actor=_actor(request))
        messages.success(request, '주소 확인 필요 상태를 해제했습니다.')
    return redirect('core:member_detail', pk=pk)


@login_required
def audit_list(request):
    q = request.GET.get('q', '').strip()
    logs = AuditLog.objects.all()
    if q:
        logs = logs.filter(Q(action__icontains=q) | Q(model_name__icontains=q) | Q(reason__icontains=q))
    return render(request, 'core/audit_list.html', {'logs': logs[:2000], 'q': q})

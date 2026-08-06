from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from django.utils import timezone

from core.models import (
    AccountType, BankTransaction, CardTransaction, Charge, ClosureEvent, ImportIssue, Member,
    MonthlyJob, ParsedRow, Payment, UploadedFile, Vehicle,
)
from core.services.audit import log_action
from core.utils import (
    extract_purpose_char, normalize_header, normalize_text, normalize_vehicle_no,
    json_safe_value, parse_date, parse_datetime, parse_decimal, stable_hash,
)


HEADER_ALIASES = {
    UploadedFile.SlotType.LICENSE: {
        'name': ['성명', '이름', '대표자', '차주명', '회원명'],
        'birth6': ['주민번호앞자리', '주민등록번호앞자리', '생년월일', '주민번호', '주민등록번호'],
        'phone': ['휴대전화', '휴대폰', '핸드폰', '전화번호', '연락처', '휴대전화번호'],
        'address': ['주소', '사업장주소', '거주지주소'],
        'official_address': ['공문주소', '송달주소', '우편주소'],
        'memo': ['비고', '특이사항', '메모'],
        'region': ['지역', '시군', '시군구', '관할'],
        'join_date': ['협회가입일자', '협회가입일', '가입일자', '가입일', '협회가입'],
        'vehicle_no': ['차량번호', '자동차등록번호', '등록번호', '차번'],
        'certificate_date': ['자격증명발급일자', '자격증명발급일', '발급일자', '발급일'],
    },
    UploadedFile.SlotType.BANK_1: {
        'transaction_at': ['거래일시', '거래일자', '거래일', '일시', '일자', '입금일'],
        'payer_text': ['입금자명', '입금자', '적요', '기재내용', '거래내용', '보낸분', '의뢰인'],
        'amount': ['입금액', '입금금액', '거래금액', '금액', '맡기신금액'],
        'bank_account': ['계좌번호', '통장', '계좌', '통장명'],
        'transaction_id': ['거래번호', '거래고유번호', '거래id', '순번'],
    },
    UploadedFile.SlotType.ALTOLAN: {
        'transaction_id': ['거래번호', '승인번호', '결제번호', '주문번호'],
        'transaction_at': ['결제일시', '거래일시', '결제일', '거래일', '일자'],
        'vehicle_no': ['차량번호', '차번', '등록번호'],
        'name': ['성명', '이름', '회원명', '결제자명'],
        'gross': ['결제금액', '총결제액', '승인금액', '금액'],
        'fee': ['수수료', '카드수수료'],
        'net': ['정산금액', '순정산액', '입금예정액'],
        'settlement_date': ['정산일', '입금일', '지급일'],
    },
    UploadedFile.SlotType.CIDER: {
        'transaction_id': ['거래번호', '승인번호', '결제번호', '주문번호'],
        'transaction_at': ['결제일시', '거래일시', '결제일', '거래일', '일자'],
        'vehicle_no': ['차량번호', '차번', '등록번호'],
        'name': ['성명', '이름', '회원명', '결제자명'],
        'gross': ['결제금액', '총결제액', '승인금액', '금액'],
        'fee': ['수수료', '카드수수료'],
        'net': ['정산금액', '순정산액', '입금예정액'],
        'settlement_date': ['정산일', '입금일', '지급일'],
    },
    UploadedFile.SlotType.RECEIVABLES: {
        'region': ['지역', '시군', '시군구', '관할'],
        'account_type': ['계정', '구분', '부과구분', '계정과목'],
        'memo': ['비고', '특이사항', '메모'],
        'vehicle_no': ['차량번호', '차번', '등록번호'],
        'name': ['성명', '이름', '회원명'],
        'charge_date': ['부과일', '부과일자', '기준일', '월'],
        'amount': ['부과금', '부과액', '금액'],
        'balance': ['미수금', '잔액', '미납액'],
    },
}
# Bank slots share aliases.
HEADER_ALIASES[UploadedFile.SlotType.BANK_2] = HEADER_ALIASES[UploadedFile.SlotType.BANK_1]
HEADER_ALIASES[UploadedFile.SlotType.BANK_3] = HEADER_ALIASES[UploadedFile.SlotType.BANK_1]

REQUIRED_FIELDS = {
    UploadedFile.SlotType.LICENSE: {'name', 'vehicle_no'},
    UploadedFile.SlotType.BANK_1: {'payer_text', 'amount'},
    UploadedFile.SlotType.BANK_2: {'payer_text', 'amount'},
    UploadedFile.SlotType.BANK_3: {'payer_text', 'amount'},
    UploadedFile.SlotType.ALTOLAN: {'gross'},
    UploadedFile.SlotType.CIDER: {'gross'},
    UploadedFile.SlotType.RECEIVABLES: {'name', 'balance'},
}


def validate_excel_signature(uploaded_file):
    ext = Path(uploaded_file.name).suffix.lower()
    if ext not in {'.xls', '.xlsx', '.xlsm'}:
        raise ValueError('xls, xlsx 또는 xlsm 파일만 업로드할 수 있습니다.')
    pos = uploaded_file.tell()
    signature = uploaded_file.read(8)
    uploaded_file.seek(pos)
    if ext in {'.xlsx', '.xlsm'} and not signature.startswith(b'PK'):
        raise ValueError('확장자와 실제 Excel 파일 형식이 맞지 않습니다.')
    if ext == '.xls' and not signature.startswith(bytes.fromhex('D0CF11E0')):
        raise ValueError('확장자는 xls이지만 실제 구형 Excel 파일 형식이 아닙니다.')


def _read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    result = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        result.append((ws.title, rows))
    wb.close()
    return result


def _read_xls(path):
    import xlrd
    book = xlrd.open_workbook(path, on_demand=True)
    result = []
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            values = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, book.datemode)
                values.append(value)
            rows.append(values)
        result.append((sheet.name, rows))
    book.release_resources()
    return result


def read_workbook(path: str):
    ext = Path(path).suffix.lower()
    if ext in {'.xlsx', '.xlsm'}:
        return _read_xlsx(path)
    if ext == '.xls':
        return _read_xls(path)
    raise ValueError('지원하지 않는 파일 형식입니다.')


def _find_column_by_alias(row, aliases):
    """Return the column index for the highest-priority matching alias.

    Field alias order matters. For example, mobile-phone headers are preferred
    over a landline ``전화번호`` column when both are present.
    """
    normalized_row = [normalize_header(value) for value in row]
    for alias in aliases:
        normalized_alias = normalize_header(alias)
        for col_index, value in enumerate(normalized_row):
            if value == normalized_alias:
                return col_index
    return None


def _latest_receivable_balance_column(rows, header_index):
    """Find the latest monthly ``N월 미수금`` column that actually contains data."""
    header = rows[header_index]
    candidates = []
    for col_index, value in enumerate(header):
        normalized = normalize_header(value)
        match = re.fullmatch(r'(\d{1,2})월미수금', normalized)
        if not match:
            continue
        month = int(match.group(1))
        if not 1 <= month <= 12:
            continue
        nonempty = 0
        for row in rows[header_index + 1:]:
            cell = row[col_index] if col_index < len(row) else None
            if cell not in (None, ''):
                nonempty += 1
        candidates.append((month, nonempty, col_index))
    populated = [item for item in candidates if item[1] > 0]
    target = max(populated or candidates, default=None, key=lambda item: item[0])
    return target[2] if target else None


def detect_header(rows, slot_type: str, scan_limit: int = 60):
    aliases = HEADER_ALIASES.get(slot_type, {})
    best = None
    for index, row in enumerate(rows[:scan_limit]):
        mapping = {}
        for field, names in aliases.items():
            col_index = _find_column_by_alias(row, names)
            if col_index is not None:
                mapping[field] = col_index

        # The legacy receivables workbook labels balances as "1월 미수금",
        # "2월 미수금" ... rather than a plain "미수금" column. Select the
        # latest month with real values, ignoring future blank month blocks.
        if slot_type == UploadedFile.SlotType.RECEIVABLES:
            balance_col = _latest_receivable_balance_column(rows, index)
            if balance_col is not None:
                mapping['balance'] = balance_col

        score = len(mapping)
        if best is None or score > best[0]:
            best = (score, index, mapping)
    return best or (0, 0, {})


def _headers_with_unique_names(header_row):
    counter = Counter()
    headers = []
    for index, value in enumerate(header_row):
        base = str(value).strip() if value not in (None, '') else f'열{index + 1}'
        counter[base] += 1
        headers.append(base if counter[base] == 1 else f'{base}_{counter[base]}')
    return headers


@transaction.atomic
def parse_uploaded_file(uploaded: UploadedFile):
    uploaded.parse_error = ''
    uploaded.parsed_rows.all().delete()
    uploaded.import_issues.all().delete()
    try:
        sheets = read_workbook(uploaded.file.path)
        summary = {'sheets': [], 'rows': 0, 'skipped_sheets': []}
        detected_headers = []
        chosen_header_row = None
        best_mapping = {}
        best_score = -1
        manual_mapping = uploaded.column_mapping or {}
        required_fields = REQUIRED_FIELDS.get(uploaded.slot_type, set())

        # The real license workbook contains helper/company/summary sheets.
        # When the normal "개인" and "택배" sheets exist, only those are the
        # current one-row-per-vehicle member ledgers.
        normalized_sheet_names = {normalize_text(name) for name, _ in sheets}
        use_primary_license_sheets = (
            uploaded.slot_type == UploadedFile.SlotType.LICENSE
            and bool(normalized_sheet_names & {'개인', '택배'})
        )

        # The real receivables workbook also contains summaries and prior-year
        # copies. Prefer the current-year detailed ledger.
        preferred_receivables_sheet = ''
        if uploaded.slot_type == UploadedFile.SlotType.RECEIVABLES:
            exact = normalize_text(f'{uploaded.job.year}년회비내역')
            if exact in normalized_sheet_names:
                preferred_receivables_sheet = exact
            else:
                preferred_receivables_sheet = next(
                    (name for name in normalized_sheet_names if '회비내역' in name),
                    '',
                )

        valid_sheet_count = 0
        for sheet_name, rows in sheets:
            if not rows:
                continue
            normalized_sheet_name = normalize_text(sheet_name)
            if use_primary_license_sheets and normalized_sheet_name not in {'개인', '택배'}:
                summary['skipped_sheets'].append({
                    'name': sheet_name, 'reason': '보조·업체·집계 시트 제외',
                })
                continue
            if preferred_receivables_sheet and normalized_sheet_name != preferred_receivables_sheet:
                summary['skipped_sheets'].append({
                    'name': sheet_name, 'reason': '현재연도 회비내역 시트가 아님',
                })
                continue

            score, header_index, detected_mapping = detect_header(rows, uploaded.slot_type)
            sheet_mapping = manual_mapping or detected_mapping
            missing = required_fields - set(sheet_mapping)
            if missing:
                summary['skipped_sheets'].append({
                    'name': sheet_name,
                    'reason': '필수 열 없음: ' + ', '.join(sorted(missing)),
                    'header_row': header_index + 1,
                    'score': score,
                })
                continue

            valid_sheet_count += 1
            headers = _headers_with_unique_names(rows[header_index])
            if score > best_score:
                best_score = score
                best_mapping = dict(sheet_mapping)
                detected_headers = headers
                chosen_header_row = header_index + 1

            row_count = 0
            for source_index, values in enumerate(rows[header_index + 1:], start=header_index + 2):
                if not any(v not in (None, '') for v in values):
                    continue
                raw = {
                    headers[i]: json_safe_value(values[i] if i < len(values) else None)
                    for i in range(len(headers))
                }
                canonical = {}
                for field, mapped in sheet_mapping.items():
                    if isinstance(mapped, int):
                        value = values[mapped] if mapped < len(values) else None
                    else:
                        value = raw.get(str(mapped))
                    canonical[field] = json_safe_value(value)
                raw['__canonical__'] = canonical
                ParsedRow.objects.create(
                    uploaded_file=uploaded,
                    sheet_name=sheet_name,
                    source_row=source_index,
                    raw_data=raw,
                    row_hash=stable_hash({
                        'sheet': sheet_name,
                        'row': source_index,
                        'canonical': canonical,
                        'raw': raw,
                    }),
                )
                row_count += 1
            summary['sheets'].append({
                'name': sheet_name,
                'rows': row_count,
                'header_row': header_index + 1,
                'score': score,
                'mapping': sheet_mapping,
            })
            summary['rows'] += row_count

        uploaded.detected_headers = detected_headers
        uploaded.header_row = chosen_header_row
        uploaded.column_mapping = best_mapping
        if valid_sheet_count == 0:
            uploaded.parse_status = UploadedFile.ParseStatus.NEEDS_MAPPING
            skipped = summary.get('skipped_sheets') or []
            detail = '; '.join(
                f"{item.get('name')}: {item.get('reason')}" for item in skipped[:6]
            )
            uploaded.parse_error = (
                '현재 원본에서 회원자료가 있는 시트를 찾지 못했습니다.'
                + (f' ({detail})' if detail else '')
            )
        else:
            uploaded.parse_status = UploadedFile.ParseStatus.PARSED
        uploaded.parse_summary = summary
        uploaded.save(update_fields=[
            'detected_headers', 'header_row', 'column_mapping', 'parse_status',
            'parse_error', 'parse_summary', 'updated_at',
        ])
        return summary
    except Exception as exc:
        uploaded.parse_status = UploadedFile.ParseStatus.FAILED
        uploaded.parse_error = f'{type(exc).__name__}: {exc}'
        uploaded.save(update_fields=['parse_status', 'parse_error', 'updated_at'])
        raise


def _mapped_value(row: ParsedRow, uploaded: UploadedFile, field: str):
    canonical = (row.raw_data or {}).get('__canonical__') or {}
    if field in canonical:
        return canonical.get(field)

    # Backward compatibility for files parsed before v3.0.3.
    mapping = uploaded.column_mapping or {}
    mapped = mapping.get(field)
    if mapped is None:
        return None
    if isinstance(mapped, int):
        headers = uploaded.detected_headers or []
        if 0 <= mapped < len(headers):
            return row.raw_data.get(headers[mapped])
        return None
    return row.raw_data.get(str(mapped))


def _sanitize_birth6(value):
    text = ''.join(ch for ch in str(value or '') if ch.isdigit())
    return text[:6] if len(text) >= 6 else ''


def _normalize_person_name(value):
    return re.sub(r'\s+', '', str(value or '')).strip()


def _normalize_phone(value):
    text = str(value or '')
    # Prefer the first mobile number when a cell contains multiple numbers or notes.
    for match in re.finditer(r'01[016789][0-9\-\s]{7,12}', text):
        digits = ''.join(ch for ch in match.group(0) if ch.isdigit())
        if len(digits) in {10, 11}:
            return digits
    # Fallback for old rows containing a single number without an 01x prefix.
    chunks = re.findall(r'[0-9][0-9\-\s]{7,14}', text)
    for chunk in chunks:
        digits = ''.join(ch for ch in chunk if ch.isdigit())
        if 9 <= len(digits) <= 11:
            return digits
    return ''


def _active_member_candidates(name, birth6):
    qs = Member.objects.filter(name=name, is_active_record=True)
    if birth6:
        qs = qs.filter(birth6=birth6)
    return qs


@transaction.atomic
def process_license_file(uploaded: UploadedFile):
    job = uploaded.job
    created = updated = vehicle_changes = issues = reopened = 0
    for row in uploaded.parsed_rows.all():
        name = _normalize_person_name(_mapped_value(row, uploaded, 'name'))
        vehicle_no = str(_mapped_value(row, uploaded, 'vehicle_no') or '').strip()
        vehicle_digits = ''.join(ch for ch in normalize_vehicle_no(vehicle_no) if ch.isdigit())
        if not name or not vehicle_no:
            continue
        if len(vehicle_digits) < 4:
            ImportIssue.objects.create(
                uploaded_file=uploaded,
                sheet_name=row.sheet_name,
                source_row=row.source_row,
                issue_type='invalid_vehicle_number',
                message='차량번호가 완전하지 않아 자동 등록하지 않았습니다.',
                raw_data=row.raw_data,
            )
            issues += 1
            continue
        birth6 = _sanitize_birth6(_mapped_value(row, uploaded, 'birth6'))
        address = str(_mapped_value(row, uploaded, 'address') or '').strip()
        official_address = str(_mapped_value(row, uploaded, 'official_address') or '').strip()
        phone = _normalize_phone(_mapped_value(row, uploaded, 'phone'))
        memo = str(_mapped_value(row, uploaded, 'memo') or '').strip()
        region = str(_mapped_value(row, uploaded, 'region') or '').strip()
        join_raw = _mapped_value(row, uploaded, 'join_date')
        join_date = parse_date(join_raw, default_year=job.year)
        cert_date = parse_date(_mapped_value(row, uploaded, 'certificate_date'), default_year=job.year)

        name_candidates = list(Member.objects.filter(name=name, is_active_record=True))
        if not birth6 and name_candidates:
            ImportIssue.objects.create(
                uploaded_file=uploaded, sheet_name=row.sheet_name, source_row=row.source_row,
                issue_type='name_only_no_auto_link',
                message='주민번호 앞자리가 없어 이름만으로 기존 회원에 자동 연결하지 않았습니다.',
                candidate_member_ids=[m.id for m in name_candidates], raw_data=row.raw_data,
            )
            issues += 1
            continue
        candidates = list(_active_member_candidates(name, birth6))
        exact_address = [m for m in candidates if normalize_text(m.address) == normalize_text(address)] if address else []
        if len(exact_address) == 1:
            member = exact_address[0]
        elif len(candidates) == 1 and birth6 and address and normalize_text(candidates[0].address) != normalize_text(address):
            ImportIssue.objects.create(
                uploaded_file=uploaded, sheet_name=row.sheet_name, source_row=row.source_row,
                issue_type='same_name_birth_address_diff',
                message='성명과 주민번호 앞자리는 같지만 주소가 달라 자동 연결하지 않았습니다.',
                candidate_member_ids=[candidates[0].id], raw_data=row.raw_data,
            )
            issues += 1
            continue
        elif len(candidates) > 1:
            ImportIssue.objects.create(
                uploaded_file=uploaded, sheet_name=row.sheet_name, source_row=row.source_row,
                issue_type='duplicate_identity_candidates',
                message='동일인 후보가 여러 명입니다.',
                candidate_member_ids=[m.id for m in candidates], raw_data=row.raw_data,
            )
            issues += 1
            continue
        else:
            member = Member(name=name, birth6=birth6, first_seen_on=job.period_start)

        is_new = member.pk is None
        old_cert = member.certificate_issued_on
        was_closed = member.operational_status == Member.OperationalStatus.CLOSED

        member.phone = phone or member.phone
        member.address = address or member.address
        if official_address:
            member.official_address = official_address
            member.official_address_custom = normalize_text(official_address) != normalize_text(address)
        elif not member.official_address_custom and member.address:
            member.official_address = member.address
        member.memo = memo
        member.region = region
        memo_key = normalize_text(memo)
        member.phone_needs_check = '결번' in memo_key
        member.sms_opt_out = '수신거부' in memo_key
        member.membership_mark_raw = str(join_raw or '')
        joined = bool(join_date or normalize_text(join_raw) in {'o', '0', '○', '가입', 'y', 'yes'})
        # 화면에서 처리한 가입·탈퇴 이력이 있으면 업로드 파일이 그 상태를 덮어쓰지 않는다.
        has_system_membership_history = bool(member.pk and member.membership_events.exists())
        if not has_system_membership_history:
            if joined:
                member.membership_status = Member.MembershipStatus.ACTIVE
                member.receivable_account_type = AccountType.MEMBERSHIP_FEE
                if join_date:
                    member.membership_started_on = join_date
                elif not member.membership_billing_anchor:
                    # 날짜 없는 O 표시는 기존 가입자로 보고 해당 작업월 1일을 부과 기준으로 둔다.
                    member.membership_billing_anchor = job.period_start
            elif member.membership_status != Member.MembershipStatus.PENDING:
                member.membership_status = Member.MembershipStatus.NON_MEMBER
                member.receivable_account_type = AccountType.MANAGEMENT_FEE
        if cert_date:
            member.certificate_issued_on = cert_date
            if old_cert != cert_date:
                member.certificate_date_recorded_on = timezone.localdate()
        if not member.first_seen_on:
            member.first_seen_on = job.period_start
        if was_closed:
            reentry_date = cert_date or job.period_start
            member.operational_status = Member.OperationalStatus.ACTIVE
            member.re_registered_on = reentry_date
            ClosureEvent.objects.create(
                member=member,
                event_type=ClosureEvent.EventType.REOPEN,
                effective_date=reentry_date,
                memo='전체면허자현황에서 동일인 재등록 자동 확인',
                actor='admin',
            )
            reopened += 1
        member.source_row_key = f'{uploaded.id}:{row.sheet_name}:{row.source_row}'
        member.save()
        if is_new:
            created += 1
        else:
            updated += 1

        normalized = normalize_vehicle_no(vehicle_no)
        current = member.current_vehicle
        if not current or current.normalized_vehicle_no != normalized:
            if current:
                current.is_current = False
                current.end_date = job.period_start
                current.change_reason = '전체면허자현황 갱신'
                current.save(update_fields=['is_current', 'end_date', 'change_reason', 'updated_at'])
            Vehicle.objects.create(
                member=member,
                vehicle_no=vehicle_no,
                normalized_vehicle_no=normalized,
                purpose_char=extract_purpose_char(vehicle_no),
                start_date=job.period_start,
                is_current=True,
                change_reason='전체면허자현황 갱신',
            )
            vehicle_changes += 1

    uploaded.parse_status = UploadedFile.ParseStatus.PROCESSED
    uploaded.parse_summary = {
        **(uploaded.parse_summary or {}), 'created_members': created, 'updated_members': updated,
        'vehicle_changes': vehicle_changes, 'reopened_members': reopened, 'issues': issues,
    }
    uploaded.save(update_fields=['parse_status', 'parse_summary', 'updated_at'])
    log_action(action='process_license_file', instance=uploaded, after=uploaded.parse_summary)
    return uploaded.parse_summary


def _bank_label(uploaded):
    return uploaded.get_slot_type_display()


@transaction.atomic
def process_bank_file(uploaded: UploadedFile):
    from core.services.ledger import replace_payment_allocations

    job = uploaded.job
    created = skipped = changed_count = 0
    for row in uploaded.parsed_rows.all():
        amount = parse_decimal(_mapped_value(row, uploaded, 'amount'))
        if amount is None or amount <= 0:
            skipped += 1
            continue
        transaction_at = parse_datetime(_mapped_value(row, uploaded, 'transaction_at'), default_year=job.year)
        payer = str(_mapped_value(row, uploaded, 'payer_text') or '').strip()
        bank_account = str(_mapped_value(row, uploaded, 'bank_account') or _bank_label(uploaded)).strip()
        source_txn_id = str(_mapped_value(row, uploaded, 'transaction_id') or '').strip()
        fingerprint = stable_hash({
            'dt': str(transaction_at or ''), 'payer': normalize_text(payer),
            'amount': str(amount), 'bank': normalize_text(bank_account),
        })
        txn_key = source_txn_id or stable_hash({
            'file': uploaded.id, 'sheet': row.sheet_name, 'row': row.source_row,
        })
        tx, was_created = BankTransaction.objects.get_or_create(
            uploaded_file=uploaded,
            source_sheet=row.sheet_name,
            source_row=row.source_row,
            defaults={
                'job': job, 'txn_key': txn_key, 'transaction_at': transaction_at,
                'payer_text': payer, 'amount': amount, 'bank_account_label': bank_account,
                'raw_data': row.raw_data, 'duplicate_group_key': fingerprint,
            },
        )
        changed = False
        if not was_created:
            changed = any([
                tx.transaction_at != transaction_at,
                tx.payer_text != payer,
                tx.amount != amount,
                tx.bank_account_label != bank_account,
                tx.txn_key != txn_key,
            ])
            tx.transaction_at = transaction_at
            tx.payer_text = payer
            tx.amount = amount
            tx.bank_account_label = bank_account
            tx.raw_data = row.raw_data
            tx.duplicate_group_key = fingerprint
            tx.txn_key = txn_key
            if changed:
                tx.status = BankTransaction.Status.REVIEW
                tx.match_reason = '원본 거래 변경으로 기존 배정 취소'
                changed_count += 1
            tx.save()
        if not hasattr(tx, 'payment'):
            payment = Payment.objects.create(
                source_type=Payment.SourceType.BANK,
                payment_date=transaction_at or timezone.now(),
                amount=amount,
                bank_transaction=tx,
                monthly_job=job,
            )
        else:
            payment = tx.payment
            payment.payment_date = transaction_at or payment.payment_date
            payment.amount = amount
            payment.save(update_fields=['payment_date', 'amount', 'updated_at'])
            if changed and payment.allocation_lines.filter(status='active').exists():
                replace_payment_allocations(payment, [], reason='원본 통장거래 변경으로 자동 원상복구')
        created += int(was_created)

    # 이전 중복표시는 재평가하되 수동배정된 건은 보존한다.
    for tx in BankTransaction.objects.filter(job=job, status=BankTransaction.Status.DUPLICATE, is_effective=True):
        if tx.allocated_amount > 0:
            tx.status = BankTransaction.Status.MANUAL_MATCHED
        else:
            tx.status = BankTransaction.Status.UNMATCHED
        tx.save(update_fields=['status', 'updated_at'])

    # 파일 간 동일거래는 삭제하지 않고 중복의심으로 표시한다.
    groups = defaultdict(list)
    for tx in BankTransaction.objects.filter(job=job, is_effective=True):
        groups[tx.duplicate_group_key].append(tx)
    for group, rows in groups.items():
        file_ids = {r.uploaded_file_id for r in rows}
        if group and len(rows) > 1 and len(file_ids) > 1:
            for tx in rows:
                if tx.status != BankTransaction.Status.MANUAL_MATCHED:
                    tx.status = BankTransaction.Status.DUPLICATE
                    tx.save(update_fields=['status', 'updated_at'])

    uploaded.parse_status = UploadedFile.ParseStatus.PROCESSED
    uploaded.parse_summary = {
        **(uploaded.parse_summary or {}), 'created_transactions': created,
        'changed_transactions': changed_count, 'skipped': skipped,
    }
    uploaded.save(update_fields=['parse_status', 'parse_summary', 'updated_at'])
    log_action(action='process_bank_file', instance=uploaded, after=uploaded.parse_summary)
    return uploaded.parse_summary


@transaction.atomic
def process_card_file(uploaded: UploadedFile):
    from core.services.ledger import replace_payment_allocations

    job = uploaded.job
    provider = CardTransaction.Provider.ALTOLAN if uploaded.slot_type == UploadedFile.SlotType.ALTOLAN else CardTransaction.Provider.CIDER
    created = skipped = duplicates = changed_count = 0
    seen_keys = set()
    for row in uploaded.parsed_rows.all():
        gross = parse_decimal(_mapped_value(row, uploaded, 'gross'))
        if gross is None or gross <= 0:
            skipped += 1
            continue
        fee = parse_decimal(_mapped_value(row, uploaded, 'fee')) or Decimal('0')
        net = parse_decimal(_mapped_value(row, uploaded, 'net'))
        if net is None:
            net = gross - fee
        transaction_at = parse_datetime(_mapped_value(row, uploaded, 'transaction_at'), default_year=job.year)
        settlement_date = parse_date(_mapped_value(row, uploaded, 'settlement_date'), default_year=job.year)
        vehicle_no = str(_mapped_value(row, uploaded, 'vehicle_no') or '').strip()
        name = str(_mapped_value(row, uploaded, 'name') or '').strip()
        source_id = str(_mapped_value(row, uploaded, 'transaction_id') or '').strip()
        if provider == CardTransaction.Provider.CIDER and source_id:
            txn_key = f'cider:{source_id}'
        else:
            txn_key = stable_hash({
                'provider': provider, 'dt': str(transaction_at or ''),
                'vehicle': normalize_vehicle_no(vehicle_no), 'name': normalize_text(name),
                'gross': str(gross),
            })
        tx, was_created = CardTransaction.objects.get_or_create(
            uploaded_file=uploaded, source_sheet=row.sheet_name, source_row=row.source_row,
            defaults={
                'job': job, 'provider': provider, 'txn_key': txn_key,
                'transaction_at': transaction_at, 'vehicle_no': vehicle_no, 'member_name': name,
                'gross_amount': gross, 'fee_amount': fee, 'net_amount': net,
                'settlement_date': settlement_date, 'raw_data': row.raw_data,
            },
        )
        changed = False
        if not was_created:
            changed = any([
                tx.txn_key != txn_key, tx.transaction_at != transaction_at,
                tx.vehicle_no != vehicle_no, tx.member_name != name,
                tx.gross_amount != gross, tx.fee_amount != fee, tx.net_amount != net,
                tx.settlement_date != settlement_date,
            ])
            tx.txn_key = txn_key
            tx.transaction_at = transaction_at
            tx.vehicle_no = vehicle_no
            tx.member_name = name
            tx.gross_amount = gross
            tx.fee_amount = fee
            tx.net_amount = net
            tx.settlement_date = settlement_date
            tx.raw_data = row.raw_data
            if changed:
                tx.status = CardTransaction.Status.REVIEW
                changed_count += 1
            tx.save()

        duplicate = (
            txn_key in seen_keys
            or CardTransaction.objects.filter(job=job, provider=provider, txn_key=txn_key)
                .exclude(pk=tx.pk).exists()
        )
        seen_keys.add(txn_key)
        if duplicate and tx.status != CardTransaction.Status.MATCHED:
            tx.status = CardTransaction.Status.DUPLICATE
            tx.duplicate_suspected = True
            tx.save(update_fields=['status', 'duplicate_suspected', 'updated_at'])
        elif not duplicate and tx.duplicate_suspected and tx.status != CardTransaction.Status.MATCHED:
            tx.duplicate_suspected = False
            tx.status = CardTransaction.Status.REVIEW if changed else CardTransaction.Status.UNMATCHED
            tx.save(update_fields=['status', 'duplicate_suspected', 'updated_at'])

        if not hasattr(tx, 'payment'):
            payment = Payment.objects.create(
                source_type=Payment.SourceType.CARD,
                payment_date=transaction_at or timezone.now(),
                amount=gross,
                card_transaction=tx,
                monthly_job=job,
            )
        else:
            payment = tx.payment
            payment.payment_date = transaction_at or payment.payment_date
            payment.amount = gross
            payment.save(update_fields=['payment_date', 'amount', 'updated_at'])
            if changed and payment.allocation_lines.filter(status='active').exists():
                replace_payment_allocations(payment, [], reason='원본 카드거래 변경으로 자동 원상복구')
        duplicates += int(duplicate)
        created += int(was_created)

    uploaded.parse_status = UploadedFile.ParseStatus.PROCESSED
    uploaded.parse_summary = {
        **(uploaded.parse_summary or {}), 'created_transactions': created,
        'changed_transactions': changed_count, 'duplicates': duplicates, 'skipped': skipped,
    }
    uploaded.save(update_fields=['parse_status', 'parse_summary', 'updated_at'])
    log_action(action='process_card_file', instance=uploaded, after=uploaded.parse_summary)
    return uploaded.parse_summary


def _account_type_from_value(value):
    normalized = normalize_text(value)
    if '협회' in normalized:
        return AccountType.MEMBERSHIP_FEE
    if '관리' in normalized:
        return AccountType.MANAGEMENT_FEE
    return None


@transaction.atomic
def process_receivables_file(uploaded: UploadedFile):
    """Import current balances, account labels, and receivables remarks."""
    created = updated = prepayments = issues = 0
    job = uploaded.job

    members = list(
        Member.objects.filter(is_active_record=True)
        .prefetch_related('vehicles')
    )
    by_exact_vehicle = defaultdict(list)
    by_name_suffix = defaultdict(list)
    by_name = defaultdict(list)
    for member in members:
        normalized_name = _normalize_person_name(member.name)
        by_name[normalized_name].append(member)
        for vehicle in member.vehicles.all():
            normalized_vehicle = vehicle.normalized_vehicle_no or normalize_vehicle_no(vehicle.vehicle_no)
            if not normalized_vehicle:
                continue
            by_exact_vehicle[normalized_vehicle].append(member)
            digits = ''.join(ch for ch in normalized_vehicle if ch.isdigit())
            if len(digits) >= 4:
                by_name_suffix[(normalized_name, digits[-4:])].append(member)

    def unique(items):
        result = []
        seen = set()
        for item in items:
            if item.id not in seen:
                seen.add(item.id)
                result.append(item)
        return result

    for row in uploaded.parsed_rows.all():
        name = _normalize_person_name(_mapped_value(row, uploaded, 'name'))
        balance = parse_decimal(_mapped_value(row, uploaded, 'balance'))
        if not name or balance is None:
            continue

        vehicle_raw = _mapped_value(row, uploaded, 'vehicle_no')
        vehicle_no = normalize_vehicle_no(vehicle_raw)
        vehicle_digits = ''.join(ch for ch in vehicle_no if ch.isdigit())
        suffix = vehicle_digits[-4:] if len(vehicle_digits) >= 4 else ''

        candidates = unique(by_exact_vehicle.get(vehicle_no, [])) if vehicle_no else []
        match_reason = '차량번호 정확히 일치'
        if len(candidates) != 1 and suffix:
            candidates = unique(by_name_suffix.get((name, suffix), []))
            match_reason = '이름과 차량 끝번호 일치'
        if len(candidates) != 1:
            candidates = unique(by_name.get(name, []))
            match_reason = '고유한 이름 일치'

        if len(candidates) != 1:
            ImportIssue.objects.create(
                uploaded_file=uploaded, sheet_name=row.sheet_name, source_row=row.source_row,
                issue_type='opening_receivable_member_match',
                message='기초 미수금 회원을 하나로 확정할 수 없습니다.',
                candidate_member_ids=[member.id for member in candidates],
                raw_data={
                    **row.raw_data,
                    '__match_attempt__': {
                        'name': name,
                        'vehicle': str(vehicle_raw or ''),
                        'vehicle_suffix': suffix,
                        'candidate_count': len(candidates),
                    },
                },
            )
            issues += 1
            continue

        member = candidates[0]
        region = str(_mapped_value(row, uploaded, 'region') or '').strip()
        memo = str(_mapped_value(row, uploaded, 'memo') or '').strip()
        account_type = _account_type_from_value(_mapped_value(row, uploaded, 'account_type'))
        if not account_type:
            account_type = (
                AccountType.MEMBERSHIP_FEE
                if member.membership_status == Member.MembershipStatus.ACTIVE
                else AccountType.MANAGEMENT_FEE
            )

        changed = False
        if region and member.region != region:
            member.region = region
            changed = True
        if member.receivable_account_type != account_type:
            member.receivable_account_type = account_type
            changed = True
        if memo:
            existing_memo = member.memo or ''
            if normalize_text(memo) not in normalize_text(existing_memo):
                member.memo = '\n'.join(part for part in [existing_memo, memo] if part).strip()
                changed = True
        if changed:
            member.save()

        if balance < 0:
            from core.services.ledger import replace_payment_allocations

            opening_key = (
                f'기초 선납금:{uploaded.id}:{row.sheet_name}:{row.source_row}'
            )
            payment = Payment.objects.filter(
                source_type=Payment.SourceType.OPENING,
                memo=opening_key,
                is_effective=True,
            ).first()
            payment_date = timezone.make_aware(
                datetime.combine(job.period_start, datetime.min.time()),
                timezone.get_current_timezone(),
            )
            if payment is None:
                payment = Payment.objects.create(
                    source_type=Payment.SourceType.OPENING,
                    payment_date=payment_date,
                    amount=abs(balance),
                    monthly_job=None,
                    memo=opening_key,
                )
            else:
                payment.payment_date = payment_date
                payment.amount = abs(balance)
                payment.save(update_fields=['payment_date', 'amount', 'updated_at'])
            replace_payment_allocations(
                payment,
                [{
                    'member': member,
                    'account_type': account_type,
                    'amount': abs(balance),
                    'memo': '기존 미수금 파일의 음수잔액(선납금)',
                }],
                reason='기초 선납금 이관',
                actor='admin',
            )
            prepayments += 1
            continue
        if balance == 0:
            updated += 1
            continue

        charge_date = job.period_start
        _, was_created = Charge.objects.update_or_create(
            member=member,
            account_type=account_type,
            charge_date=charge_date,
            monthly_job=None,
            defaults={
                'amount': balance,
                'source_rule': f'opening_receivable:{match_reason}',
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    uploaded.parse_status = UploadedFile.ParseStatus.PROCESSED
    uploaded.parse_summary = {
        **(uploaded.parse_summary or {}),
        'opening_charges': created,
        'opening_charges_updated': updated,
        'opening_prepayments': prepayments,
        'issues': issues,
    }
    uploaded.save(update_fields=['parse_status', 'parse_summary', 'updated_at'])
    log_action(action='process_receivables_file', instance=uploaded, after=uploaded.parse_summary)
    return uploaded.parse_summary


def process_uploaded_file(uploaded: UploadedFile):
    if uploaded.parse_status not in {UploadedFile.ParseStatus.PARSED, UploadedFile.ParseStatus.PROCESSED}:
        raise ValueError('먼저 열 매핑을 완료하고 파싱해야 합니다.')
    if uploaded.slot_type == UploadedFile.SlotType.LICENSE:
        return process_license_file(uploaded)
    if uploaded.slot_type in {UploadedFile.SlotType.BANK_1, UploadedFile.SlotType.BANK_2, UploadedFile.SlotType.BANK_3}:
        return process_bank_file(uploaded)
    if uploaded.slot_type in {UploadedFile.SlotType.ALTOLAN, UploadedFile.SlotType.CIDER}:
        return process_card_file(uploaded)
    if uploaded.slot_type == UploadedFile.SlotType.RECEIVABLES:
        return process_receivables_file(uploaded)
    raise ValueError('처리기가 없는 파일 유형입니다.')
